from django.shortcuts import render, redirect
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from .permissions import role_required, is_any_role, is_role
from .views_extra import course_catalog, my_certificates
from .views_profile import update_profile_photo
from .api_views import get_live_updates
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Count, Avg, Min, Max, Sum, Q, F
from django.core.cache import cache
from django.views.decorators.http import require_http_methods
from .models import Profile, Course, Attendance, Assignment, Enrollment, StaffMember, SalaryRecord, FeePayment, StudentPayout, FinancialTransaction, StaffAttendance, StaffDailyAttendance, AssignmentAttachment
from .models import Schedule
from datetime import datetime, date, timedelta
from django.utils import timezone
from .ml_predictor import PerformancePredictor
from django.core.mail import send_mail
from django.http import JsonResponse
from django.conf import settings
from .report_generator import download_student_report, StudentReportGenerator
from .teacher_report_generator import TeacherReportGenerator
from .models import Submission, Notification
from .forms import ScheduleForm
from .models import StudyMaterial, Feedback
from django.contrib.auth.models import User
from django.http import Http404
from io import BytesIO
from django.http import HttpResponse
from django.http import JsonResponse
from django.template.loader import render_to_string
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

import csv
import io
import json
from decimal import Decimal
import uuid

# -----------------------
# Report utilities
# -----------------------
def validate_date_range(start_str, end_str):
    """Validate and parse date range parameters."""
    start_date = None
    end_date = None
    errors = []
    if start_str:
        try:
            start_date = date.fromisoformat(start_str)
        except Exception:
            errors.append(f"Invalid start date format: {start_str}")
    if end_str:
        try:
            end_date = date.fromisoformat(end_str)
        except Exception:
            errors.append(f"Invalid end date format: {end_str}")
    if start_date and end_date and start_date > end_date:
        errors.append("Start date must be before end date")
    return start_date, end_date, errors


def is_admin_role(profile):
    """Return True if profile has elevated admin privileges (superadmin or admin2)."""
    return bool(profile and getattr(profile, 'role', None) in ('superadmin', 'admin2'))


def check_course_access(user, course_id: int):
    """Check if user has access to course. Returns (course, error_message)."""
    profile = getattr(user, 'profile', None)
    if not profile:
        return None, 'User profile not found'
    try:
        course = Course.objects.select_related('teacher__user').get(id=course_id)
    except Course.DoesNotExist:
        return None, 'Course not found'

    return course, None


# Public pages (minimal implementations)
def home(request):
    """Render home page with a sample of courses."""
    courses = Course.objects.all()[:12]
    return render(request, 'home.html', {'courses': courses})


def courses_full(request):
    """Render a full-page courses listing with search and filters."""
    try:
        courses = Course.objects.select_related('teacher__user').all().order_by('name')
    except Exception:
        courses = Course.objects.all()

    # Simple search support
    q = request.GET.get('q', '').strip()
    if q:
        courses = courses.filter(Q(name__icontains=q) | Q(code__icontains=q) | Q(description__icontains=q))

    # Pagination (12 per page)
    paginator = Paginator(courses, 12)
    page = request.GET.get('page')
    try:
        courses_page = paginator.page(page)
    except PageNotAnInteger:
        courses_page = paginator.page(1)
    except EmptyPage:
        courses_page = paginator.page(paginator.num_pages)

    return render(request, 'courses/full_courses.html', {
        'courses': courses_page,
        'query': q
    })


def about(request):
    return render(request, 'about.html')


def choose(request):
    return render(request, 'choose.html')


def python_full_stack(request):
    return render(request, 'python_full-stack.html')


def instructor_profile(request, slug):
    """Render a lightweight instructor profile page.

    This supports two modes:
    - If a User with username==slug exists and has a Profile, show that profile.
    - Otherwise fall back to a small static mapping for a few demo slugs.
    """
    # Try to resolve by username first
    profile = None
    try:
        user = User.objects.filter(username=slug).first()
        if user and hasattr(user, 'profile'):
            profile = user.profile
    except Exception:
        profile = None

    if profile:
        return render(request, 'instructor_profile.html', {'profile': profile})

    # Fallback static data for a couple of demo slugs used in templates
    instructors = {
        'prakash-m': {
            'name': 'Nataraj Adithya S',
            'role': 'Lead Instructor — Full-Stack & AI',
            'bio': 'Senior full-stack developer with 10+ years of experience building Django and React applications. Specializes in scalable backend systems and ML integrations.',
            'skills': 'Django, REST API, React, PostgreSQL, Docker, ML',
            'email': 'adithyaxploreitcorp@gmail.com',
            'experience': '10+',
            'photo_url': 'https://media.licdn.com/dms/image/v2/D5622AQFKYbp000NqHQ/feedshare-shrink_1280/B56ZiyBSGCHUAk-/0/1755333338175?e=1762992000&v=beta&t=hofQ3y__j0NzwfVIeiJm73S3pUjoQdP_DiJRL17hCZQ'
        },
        'nataraj-adithya-s': {
            'name': 'Nataraj Adithya S',
            'role': 'Lead Instructor — Full-Stack & AI',
            'bio': 'Senior full-stack developer with 10+ years of experience building Django and React applications. Specializes in scalable backend systems and ML integrations.',
            'skills': 'Django, REST API, React, PostgreSQL, Docker, ML',
            'email': 'adithyaxploreitcorp@gmail.com',
            'experience': '10+',
            'photo_url': 'https://media.licdn.com/dms/image/v2/D5622AQFKYbp000NqHQ/feedshare-shrink_1280/B56ZiyBSGCHUAk-/0/1755333338175?e=1762992000&v=beta&t=hofQ3y__j0NzwfVIeiJm73S3pUjoQdP_DiJRL17hCZQ'
        },
        'anjali-kapoor': {
            'name': 'Gokul',
            'role': 'Backend Instructor',
            'bio': 'Expert in backend architectures, database design, and performance tuning. Loves teaching clean, testable code and API design.',
            'skills': 'Django, PostgreSQL, Celery, Redis',
            'email': 'GOKUL@xploreitcorp.com',
            'experience': '8+',
            'photo_url': 'https://th.bing.com/th/id/OIP.FTHq_r1ZW67TsmTpaPdg4gHaHa?w=155&h=180&c=7&r=0&o=7&cb=12&dpr=1.4&pid=1.7&rm=3'
        }
    }

    info = instructors.get(slug)
    if info:
        return render(request, 'instructor_profile.html', info)

    # If nothing matches, show a friendly 404-style page
    raise Http404('Instructor not found')


@role_required('student')
def student_prediction(request):
    """Show prediction page for a student using the PerformancePredictor."""
    profile = getattr(request.user, 'profile', None)
    if not profile or getattr(profile, 'role', None) != 'student':
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    predictor = PerformancePredictor()
    try:
        pred = predictor.predict_performance(profile)
    except Exception:
        pred = None

    return render(request, 'student_prediction.html', {'prediction': pred})


def log_report_generation(user, report_type, course_id=None, params=None):
    """Best-effort logging of report downloads to ReportLog model if available."""
    try:
        from .models import ReportLog
        ReportLog.objects.create(user=user, report_type=report_type, course_id=course_id, params=json.dumps(params or {}))
    except Exception:
        # If model doesn't exist or any error occurs, silently ignore to avoid breaking downloads
        pass


@role_required(['superadmin', 'admin2'])
def admin_page(request):
    """A simple admin-facing page (not Django admin). Requires superadmin role."""
    if is_admin_role(getattr(request.user, 'profile', None)):
        return render(request, 'admin_page.html')
    else:
        from django.contrib import messages
        messages.error(request, 'Access denied!')
        # If not authorized, send to login (or role_redirect if you prefer)
        return redirect('login')


@login_required
def export_users_pdf(request):
    """Export a PDF summary of recent users. Only superadmin."""
    if not is_admin_role(getattr(request.user, 'profile', None)):
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    users = User.objects.order_by('-date_joined')[:100]
    # If ReportLab isn't installed, fall back to a CSV export which is widely supported.
    import csv
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="recent_users.csv"'

    writer = csv.writer(response)
    writer.writerow(['Username', 'Email', 'Role', 'Department', 'Joined'])
    for u in users:
        role = getattr(u, 'profile', None) and getattr(u.profile, 'role', '')
        dept = getattr(u, 'profile', None) and getattr(u.profile, 'department', '')
        writer.writerow([u.username, u.email, role or '', dept or '', u.date_joined.strftime('%Y-%m-%d')])

    return response


@login_required
def download_student_report_admin(request, user_id):
    """Allow admin to download a student's report by user id."""
    if not is_admin_role(getattr(request.user, 'profile', None)):
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        messages.error(request, 'User not found')
        return redirect('admin_dashboard')

    # Use the existing StudentReportGenerator if profile exists
    if not hasattr(user, 'profile'):
        messages.error(request, 'User has no profile')
        return redirect('admin_dashboard')

    generator = StudentReportGenerator(user.profile)
    pdf = generator.generate_report()
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="report_{user.username}.pdf"'
    response.write(pdf)
    return response


@login_required
def download_teacher_report(request):
    """Generate and download teacher performance report."""
    # Verify user has teacher role
    profile = getattr(request.user, 'profile', None)
    if not profile or getattr(profile, 'role', None) != 'teacher':
        messages.error(request, 'Access denied! This feature is only available to teachers.')
        return redirect('role_redirect')
    
    try:
        # Generate report
        generator = TeacherReportGenerator(profile, request=request)
        pdf = generator.generate_report()
        
        # Create response
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"teacher_report_{request.user.username}_{timestamp}.pdf"
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(pdf)
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error generating report: {str(e)}')
        return redirect('teacher_dashboard')


# Admin AI insights
@login_required
def admin_ai_insights(request):
    if not is_admin_role(getattr(request.user, 'profile', None)):
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    predictor = PerformancePredictor()
    predictor.train_model()
    students = Profile.objects.filter(role='student')
    predictions = []
    for student in students:
        pred = predictor.predict_performance(student)
        if pred:
            predictions.append({'student': student, 'prediction': pred})

    predictions.sort(key=lambda x: x['prediction']['score'], reverse=True)
    # Expose whether the predictor used a trained model or the heuristic fallback
    model_status = 'trained' if getattr(predictor, 'trained', False) else 'heuristic'
    return render(request, 'dashboards/admin_ai_insights.html', {'predictions': predictions, 'model_status': model_status})


# Attendance views
@role_required('teacher')
def take_attendance(request, course_id):

    course = Course.objects.get(id=course_id, teacher=request.user.profile)
    enrollments = Enrollment.objects.filter(course=course)

    if request.method == 'POST':
        attendance_date = date.today()
        for enrollment in enrollments:
            student = enrollment.student
            status = request.POST.get(f'student_{student.id}') == 'present'
            Attendance.objects.update_or_create(
                student=student,
                course=course,
                date=attendance_date,
                defaults={'status': status}
            )
        messages.success(request, 'Attendance marked successfully!')
        return redirect('teacher_dashboard')

    return render(request, 'attendance/take_attendance.html', {'course': course, 'enrollments': enrollments, 'today': date.today()})


@role_required('teacher')
def upload_schedule(request):
    """Teacher-facing schedule/event upload page."""
    profile = getattr(request.user, 'profile', None)
    if request.method == 'POST':
        form = ScheduleForm(request.POST, request.FILES)
        if form.is_valid():
            sched = form.save(commit=False)
            sched.created_by = profile
            # If teacher selected a course, ensure they own it (defensive)
            if sched.course and sched.course.teacher != profile:
                messages.error(request, 'You can only add schedules for your own courses.')
                return redirect('upload_schedule')
            # ALWAYS enable notifications when creating schedules
            sched.notify_students = True
            sched.is_active = True
            sched.save()
            messages.success(request, 'Schedule created successfully and notifications sent!')
            return redirect('teacher_schedules')
    else:
        form = ScheduleForm()

    # Provide only courses taught by this teacher for selection convenience
    courses = Course.objects.filter(teacher=profile)
    return render(request, 'schedules/upload_schedule.html', {'form': form, 'courses': courses})


@role_required('teacher')
def teacher_schedules(request):
    profile = getattr(request.user, 'profile', None)
    schedules = Schedule.objects.filter(created_by=profile).order_by('-date', 'start_time')
    return render(request, 'schedules/teacher_schedules.html', {'schedules': schedules})


@role_required('teacher')
def edit_schedule(request, schedule_id):
    """Edit an existing schedule."""
    profile = getattr(request.user, 'profile', None)
    try:
        schedule = Schedule.objects.get(id=schedule_id, created_by=profile)
    except Schedule.DoesNotExist:
        messages.error(request, 'Schedule not found or you do not have permission to edit it.')
        return redirect('teacher_schedules')
    
    if request.method == 'POST':
        form = ScheduleForm(request.POST, request.FILES, instance=schedule)
        if form.is_valid():
            sched = form.save(commit=False)
            # Verify teacher owns the selected course
            if sched.course and sched.course.teacher != profile:
                messages.error(request, 'You can only edit schedules for your own courses.')
                return redirect('teacher_schedules')
            sched.save()
            messages.success(request, 'Schedule updated successfully!')
            return redirect('teacher_schedules')
    else:
        form = ScheduleForm(instance=schedule)
    
    courses = Course.objects.filter(teacher=profile)
    return render(request, 'schedules/edit_schedule.html', {'form': form, 'schedule': schedule, 'courses': courses})


@role_required('teacher')
def delete_schedule(request, schedule_id):
    """Delete a schedule."""
    profile = getattr(request.user, 'profile', None)
    try:
        schedule = Schedule.objects.get(id=schedule_id, created_by=profile)
    except Schedule.DoesNotExist:
        messages.error(request, 'Schedule not found or you do not have permission to delete it.')
        return redirect('teacher_schedules')
    
    if request.method == 'POST':
        title = schedule.title
        schedule.delete()
        messages.success(request, f'Schedule "{title}" deleted successfully!')
        return redirect('teacher_schedules')
    
    # GET request - show confirmation page
    return render(request, 'schedules/delete_schedule.html', {'schedule': schedule})


@role_required('student')
def student_schedules(request):
    """Student-facing list of upcoming schedules/events relevant to the student."""
    profile = getattr(request.user, 'profile', None)
    # Schedules that are public OR tied to one of the student's courses
    enrollments = Enrollment.objects.filter(student=profile).values_list('course_id', flat=True)
    today = date.today()
    schedules = Schedule.objects.filter(
        Q(is_public=True) | Q(course_id__in=list(enrollments)),
        date__gte=today
    ).order_by('date', 'start_time')
    return render(request, 'schedules/student_schedules.html', {'schedules': schedules})


@role_required('student')
def view_attendance(request):

    student = request.user.profile
    enrollments = Enrollment.objects.filter(student=student)
    attendance_data = []
    for enrollment in enrollments:
        total = Attendance.objects.filter(student=student, course=enrollment.course).count()
        present = Attendance.objects.filter(student=student, course=enrollment.course, status=True).count()
        percentage = (present / total * 100) if total > 0 else 0
        attendance_data.append({'course': enrollment.course, 'total': total, 'present': present, 'absent': total - present, 'percentage': round(percentage, 2)})

    return render(request, 'attendance/view_attendance.html', {'attendance_data': attendance_data})


# Notifications
@login_required
def notifications_view(request):
    notifications = Notification.objects.filter(user=request.user)
    if request.method == 'POST':
        notif_id = request.POST.get('notif_id')
        try:
            notification = Notification.objects.get(id=notif_id, user=request.user)
            notification.is_read = True
            notification.save()
        except Notification.DoesNotExist:
            messages.error(request, 'Notification not found')
        return redirect('notifications')

    context = {'notifications': notifications, 'unread_count': notifications.filter(is_read=False).count()}
    return render(request, 'notifications.html', context)


@login_required
def mark_as_read(request):
    if request.method != 'POST':
        return redirect('notifications')

    notif_id = request.POST.get('notif_id')
    try:
        notification = Notification.objects.get(id=notif_id, user=request.user)
        notification.is_read = True
        notification.save()
        messages.success(request, 'Marked as read')
    except Notification.DoesNotExist:
        messages.error(request, 'Notification not found')

    return redirect('notifications')


@login_required
def delete_notification(request):
    if request.method != 'POST':
        return redirect('notifications')

    notif_id = request.POST.get('notif_id')
    try:
        notification = Notification.objects.get(id=notif_id, user=request.user)
        notification.delete()
        messages.success(request, 'Notification deleted')
    except Notification.DoesNotExist:
        messages.error(request, 'Notification not found')

    return redirect('notifications')


def create_notification(user, title, message):
    Notification.objects.create(user=user, title=title, message=message)


@role_required(['teacher', 'admin2', 'superadmin'])
def schedule_broadcast_notifications(request):
    """Teacher/Admin page to view schedule broadcast notifications.
    
    Shows real-time status of which students/teachers received schedule notifications.
    Data is fetched via AJAX from the get_schedule_notifications API endpoint.
    """
    return render(request, 'schedules/broadcast_notifications.html', {})



@login_required
def mark_all_read(request):
    """Mark all unread notifications for the current user as read.

    Accepts POST only (form in template uses POST). Redirects back to notifications.
    """
    if request.method != 'POST':
        return redirect('notifications')

    try:
        Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
        messages.success(request, 'All notifications marked as read')
    except Exception:
        messages.error(request, 'Could not mark notifications as read')

    return redirect('notifications')


# Registration View
def register_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        password2 = request.POST.get('password2')
        role = request.POST.get('role')
        department = request.POST.get('department', '')
        
        if password != password2:
            messages.error(request, 'Passwords do not match!')
            return redirect('register')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists!')
            return redirect('register')
        
        # Create user
        user = User.objects.create_user(
            username=username,
            email=email,
            password=password
        )

        # Update profile
        profile = user.profile
        profile.role = role
        profile.department = department

        # Handle uploaded profile photo if present
        try:
            pic = request.FILES.get('profile_pic')
            if pic:
                profile.profile_pic = pic
        except Exception:
            # non-fatal; continue without photo
            pass

        profile.save()

        messages.success(request, 'Account created successfully! Please login.')
        return redirect('login')
    
    return render(request, 'register.html')


# Login View
def login_view(request):
    if request.user.is_authenticated:
        return redirect('role_redirect')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            return redirect('role_redirect')
        else:
            messages.error(request, 'Invalid credentials!')
    
    return render(request, 'login.html')


# Logout View
def logout_view(request):
    logout(request)
    messages.success(request, 'Logged out successfully!')
    return redirect('login')


# Role-based Redirect
@login_required
def role_redirect(request):
    # Safely get the role attribute; handle missing/invalid roles to avoid redirect loops
    role = getattr(request.user.profile, 'role', None)

    # Route superadmin and admin2 to their respective dashboards
    if role == 'superadmin':
        return redirect('admin_dashboard')
    if role == 'admin2':
        return redirect('admin2_dashboard')
    elif role == 'teacher':
        return redirect('teacher_dashboard')
    elif role == 'student':
        return redirect('student_dashboard')
    else:
        # Role is missing or unrecognized — log the user out to break potential redirect loops,
        # then send them to login with a helpful message so they can re-authenticate or contact admin.
        from django.contrib import messages
        from django.contrib.auth import logout as auth_logout
        auth_logout(request)
        messages.error(request, 'Your account role is not set. You have been logged out. Please contact an administrator to set your role.')
        return redirect('login')


# Admin Dashboard
@login_required
def admin_dashboard(request):
    # Allow either a Profile with elevated admin role (superadmin or admin2)
    user_profile = getattr(request.user, 'profile', None)
    if not is_admin_role(user_profile):
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')
    # compute core stats
    # Use both User and Profile models so we correctly count users even if some
    # profiles are missing or roles are not set.
    total_users = User.objects.count()
    total_courses = Course.objects.count()

    # Profiles with explicit roles
    total_students = Profile.objects.filter(role='student').count()
    total_teachers = Profile.objects.filter(role='teacher').count()

    # Users that may not have a Profile or have no role set should be counted as 'other'
    total_profiles = Profile.objects.count()
    users_without_profile = max(0, total_users - total_profiles)

    other_count = max(0, total_users - (total_students + total_teachers))

    # recent registrations for the last 7 days
    today = timezone.now().date()
    signup_labels = []
    signup_counts = []
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        signup_labels.append(d.strftime('%b %d'))
        signup_counts.append(User.objects.filter(date_joined__date=d).count())

    context = {
        'total_students': total_students,
        'total_teachers': total_teachers,
        'total_courses': total_courses,
        'recent_users': User.objects.order_by('-date_joined')[:5],
        'other_count': other_count,
        # helper list for charts (safe to render as JSON)
        'user_counts': [total_students, total_teachers, other_count],
        'signup_labels': signup_labels,
        'signup_counts': signup_counts,
    }
    return render(request, 'dashboards/superadmin_dashboard.html', context)


@login_required
def admin2_dashboard(request):
    """Unified admin2 panel (powerful sub-admin) - shows students, staff and finance summary."""
    user_profile = getattr(request.user, 'profile', None)
    if not is_admin_role(user_profile):
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    total_students = Profile.objects.filter(role='student').count()
    total_teachers = Profile.objects.filter(role='teacher').count()
    total_staff = StaffMember.objects.count()
    total_fee_revenue = FeePayment.objects.aggregate(total=Sum('amount'))['total'] or 0

    context = {
        'total_students': total_students,
        'total_teachers': total_teachers,
        'total_staff': total_staff,
        'total_fee_revenue': total_fee_revenue,
    }
    return render(request, 'dashboards/admin2_dashboard.html', context)


@login_required
def admin2_updates(request):
    """Return a small JSON payload for the admin2 dashboard live updates."""
    user_profile = getattr(request.user, 'profile', None)
    if not is_admin_role(user_profile):
        return JsonResponse({'error': 'Access denied'}, status=403)

    total_students = Profile.objects.filter(role='student').count()
    total_teachers = Profile.objects.filter(role='teacher').count()
    total_staff = StaffMember.objects.count()
    total_fee_revenue = FeePayment.objects.aggregate(total=Sum('amount'))['total'] or 0

    recent_tx = FinancialTransaction.objects.order_by('-created_at')[:5]
    recent_payouts = StudentPayout.objects.order_by('-requested_on')[:5]

    data = {
        'total_students': total_students,
        'total_teachers': total_teachers,
        'total_staff': total_staff,
        'total_fee_revenue': float(total_fee_revenue),
        'recent_transactions': [{'title': t.title, 'amount': float(t.amount), 'type': t.trans_type} for t in recent_tx],
        'recent_payouts_count': recent_payouts.count(),
        'timestamp': datetime.now().isoformat(),
    }
    return JsonResponse(data)


@login_required
def admin2_financial(request):
    user_profile = getattr(request.user, 'profile', None)
    if not is_admin_role(user_profile):
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    transactions = FinancialTransaction.objects.order_by('-created_at')[:200]
    salaries = SalaryRecord.objects.order_by('-paid_on')[:50]
    payouts = StudentPayout.objects.order_by('-requested_on')[:50]
    fees = FeePayment.objects.order_by('-paid_on')[:50]

    balance = None
    if transactions.exists():
        balance = transactions.first().balance_after

    return render(request, 'dashboards/admin2_financial.html', {'transactions': transactions, 'salaries': salaries, 'payouts': payouts, 'fees': fees, 'balance': balance})


@login_required
def admin2_student_payments(request, student_id):
    """Show payments for a specific student and allow export to CSV/PDF."""
    user_profile = getattr(request.user, 'profile', None)
    if not is_admin_role(user_profile):
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    try:
        student = Profile.objects.get(id=student_id, role='student')
    except Profile.DoesNotExist:
        messages.error(request, 'Student not found')
        return redirect('admin2_financial')

    payments = FeePayment.objects.filter(student=student).order_by('-paid_on')

    # Totals
    total_paid = payments.aggregate(total=Sum('amount'))['total'] or 0

    return render(request, 'dashboards/admin2_student_payments.html', {'student': student, 'payments': payments, 'total_paid': total_paid})


@login_required
def export_student_payments_csv(request, student_id):
    user_profile = getattr(request.user, 'profile', None)
    if not is_admin_role(user_profile):
        return HttpResponse('Access denied', status=403)

    try:
        student = Profile.objects.get(id=student_id, role='student')
    except Profile.DoesNotExist:
        return HttpResponse('Student not found', status=404)

    payments = FeePayment.objects.filter(student=student).order_by('-paid_on')

    import csv
    import io
    si = io.StringIO()
    writer = csv.writer(si)
    writer.writerow(['Paid On', 'Amount', 'Method', 'Status'])
    for p in payments:
        writer.writerow([p.paid_on.strftime('%Y-%m-%d %H:%M:%S'), f"{p.amount}", p.payment_method or '', p.status or ''])

    resp = HttpResponse(si.getvalue().encode('utf-8'), content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="student_{student.user.username}_payments.csv"'
    return resp


@login_required
def export_student_payments_pdf(request, student_id):
    user_profile = getattr(request.user, 'profile', None)
    if not is_admin_role(user_profile):
        return HttpResponse('Access denied', status=403)

    try:
        student = Profile.objects.get(id=student_id, role='student')
    except Profile.DoesNotExist:
        return HttpResponse('Student not found', status=404)

    payments = FeePayment.objects.filter(student=student).order_by('-paid_on')

    # Use reportlab to build a simple PDF
    buffer = BytesIO()
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        story.append(Paragraph(f"Payment Report - {student.user.get_full_name() or student.user.username}", styles['Heading2']))
        story.append(Spacer(1, 12))

        data = [['Paid On', 'Amount', 'Method', 'Status']]
        for p in payments:
            data.append([p.paid_on.strftime('%Y-%m-%d %H:%M:%S'), str(p.amount), p.payment_method or '', p.status or ''])

        table = Table(data, colWidths=[150, 80, 120, 80])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0d6efd')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('GRID', (0,0), (-1,-1), 0.4, colors.black),
        ]))
        story.append(table)
        doc.build(story)
        pdf = buffer.getvalue()
        buffer.close()
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="student_{student.user.username}_payments.pdf"'
        return response
    except Exception:
        buffer.close()
        return HttpResponse('PDF generation failed (reportlab required)', status=500)


@login_required
def admin2_staff(request):
    user_profile = getattr(request.user, 'profile', None)
    if not is_admin_role(user_profile):
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    staff = StaffMember.objects.select_related('profile__user').all()
    # Also provide a list of teachers who are not yet added as StaffMember so admin can add them
    teacher_profiles = Profile.objects.filter(role='teacher')
    teacher_profiles = teacher_profiles.exclude(id__in=StaffMember.objects.values_list('profile_id', flat=True))

    return render(request, 'dashboards/admin2_staff.html', {'staff': staff, 'teachers': teacher_profiles})


@login_required
def admin2_add_staff(request, profile_id):
    """Confirm and add a teacher Profile as a StaffMember (admin2 action)."""
    user_profile = getattr(request.user, 'profile', None)
    if not is_admin_role(user_profile):
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    try:
        profile = Profile.objects.get(id=profile_id, role='teacher')
    except Profile.DoesNotExist:
        messages.error(request, 'Teacher profile not found')
        return redirect('admin2_staff')

    if request.method == 'POST':
        # create StaffMember if not exists
        if StaffMember.objects.filter(profile=profile).exists():
            messages.warning(request, 'This teacher is already a staff member')
            return redirect('admin2_staff')

        StaffMember.objects.create(profile=profile, position=request.POST.get('position', 'Teacher'))
        messages.success(request, f'{profile.user.username} added as staff')
        return redirect('admin2_staff')

    return render(request, 'dashboards/admin2_confirm_add_staff.html', {'profile': profile})


@login_required
def admin2_remove_staff(request, staff_id):
    """Remove a StaffMember entry. Accepts POST only and redirects back to staff list."""
    user_profile = getattr(request.user, 'profile', None)
    if not is_admin_role(user_profile):
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    if request.method != 'POST':
        messages.error(request, 'Invalid request')
        return redirect('admin2_staff')

    try:
        staff = StaffMember.objects.get(id=int(staff_id))
    except StaffMember.DoesNotExist:
        messages.error(request, 'Staff member not found')
        return redirect('admin2_staff')

    username = getattr(getattr(staff, 'profile', None), 'user', None)
    try:
        staff.delete()
        messages.success(request, f'Staff member removed')
    except Exception:
        messages.error(request, 'Could not remove staff member')

    return redirect('admin2_staff')


@login_required
def admin2_payouts(request):
    user_profile = getattr(request.user, 'profile', None)
    if not is_admin_role(user_profile):
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    payouts = StudentPayout.objects.order_by('-requested_on')[:200]
    return render(request, 'dashboards/admin2_payouts.html', {'payouts': payouts})


@login_required
def admin2_add_salary(request):
    """Add a salary record for a staff member and create a corresponding financial transaction."""
    user_profile = getattr(request.user, 'profile', None)
    if not is_admin_role(user_profile):
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    # Preselect staff if provided via GET param (from staff list 'Pay' links)
    selected_staff_id = request.GET.get('staff_id') or request.POST.get('staff_id')

    if request.method == 'POST':
        staff_id = request.POST.get('staff_id')
        amount = request.POST.get('amount')
        notes = request.POST.get('notes', '')
        is_advance = request.POST.get('is_advance') == 'on'
        try:
            staff = StaffMember.objects.get(id=int(staff_id))
            amt = Decimal(str(amount))
        except Exception:
            messages.error(request, 'Invalid input')
            return redirect('admin2_financial')

        # Record salary
        if is_advance:
            notes = (notes or '') + ' (advance)'
        SalaryRecord.objects.create(staff=staff, amount=amt, notes=notes)

        # Create financial transaction (debit)
        last = FinancialTransaction.objects.order_by('-created_at').first()
        last_balance = Decimal(last.balance_after) if last and last.balance_after is not None else Decimal('0.00')
        new_balance = last_balance - amt
        FinancialTransaction.objects.create(title=f"Salary payment to {staff.profile.user.username}", amount=amt, trans_type='debit', balance_after=new_balance)

        messages.success(request, 'Salary recorded')
        return redirect('admin2_financial')

    staff = StaffMember.objects.select_related('profile__user').all()
    return render(request, 'dashboards/admin2_add_salary.html', {'staff': staff, 'selected_staff_id': selected_staff_id})


@login_required
def admin2_record_fee(request):
    """Record an offline fee payment for a student and create a transaction (admin action)."""
    user_profile = getattr(request.user, 'profile', None)
    if not is_admin_role(user_profile):
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        amount = request.POST.get('amount')
        method = request.POST.get('method', 'offline')
        status = request.POST.get('status', 'paid')
        try:
            student = Profile.objects.get(id=int(student_id), role='student')
            amt = Decimal(str(amount))
        except Exception:
            messages.error(request, 'Invalid input')
            return redirect('admin2_financial')

        FeePayment.objects.create(student=student, amount=amt, payment_method=method, status=status)

        # Financial transaction (credit)
        last = FinancialTransaction.objects.order_by('-created_at').first()
        last_balance = Decimal(last.balance_after) if last and last.balance_after is not None else Decimal('0.00')
        new_balance = last_balance + amt
        FinancialTransaction.objects.create(title=f"Fee payment by {student.user.username}", amount=amt, trans_type='credit', balance_after=new_balance)

        messages.success(request, 'Fee recorded')
        return redirect('admin2_financial')

    students = Profile.objects.filter(role='student').select_related('user')[:200]
    return render(request, 'dashboards/admin2_record_fee.html', {'students': students})


@login_required
def admin2_export_financial_excel(request):
    """Export consolidated financial data as an Excel file (openpyxl optional) or CSV fallback."""
    user_profile = getattr(request.user, 'profile', None)
    if not is_admin_role(user_profile):
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    # Optional date filters
    start_str = request.GET.get('start')
    end_str = request.GET.get('end')
    start_date, end_date, errs = validate_date_range(start_str, end_str)

    tx_qs = FinancialTransaction.objects.all().order_by('-created_at')
    if start_date:
        tx_qs = tx_qs.filter(created_at__date__gte=start_date)
    if end_date:
        tx_qs = tx_qs.filter(created_at__date__lte=end_date)

    salaries = SalaryRecord.objects.all().order_by('-paid_on')
    fees = FeePayment.objects.all().order_by('-paid_on')
    payouts = StudentPayout.objects.all().order_by('-requested_on')

    # Prefer openpyxl if available (optional dependency). Use a guarded import
    # and a type-ignore comment to silence static analyzers when the package
    # isn't installed in the developer environment.
    use_openpyxl = False
    try:
        from openpyxl import Workbook  # type: ignore
        use_openpyxl = True
    except Exception:
        use_openpyxl = False

    if use_openpyxl:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Transactions'
        ws.append(['Created At', 'Title', 'Type', 'Amount', 'Balance After'])
        for t in tx_qs:
            ws.append([t.created_at.strftime('%Y-%m-%d %H:%M:%S'), t.title, t.trans_type, float(t.amount), float(t.balance_after) if t.balance_after is not None else ''])

        ws2 = wb.create_sheet('Salaries')
        ws2.append(['Paid On', 'Staff', 'Amount', 'Notes'])
        for s in salaries:
            ws2.append([s.paid_on.strftime('%Y-%m-%d'), s.staff.profile.user.username, float(s.amount), s.notes or ''])

        ws3 = wb.create_sheet('Fees')
        ws3.append(['Paid On', 'Student', 'Amount', 'Method', 'Status'])
        for f in fees:
            ws3.append([f.paid_on.strftime('%Y-%m-%d %H:%M:%S'), f.student.user.username if f.student and getattr(f.student, 'user', None) else '', float(f.amount), f.payment_method or '', f.status or ''])

        ws4 = wb.create_sheet('Payouts')
        ws4.append(['Requested On', 'Student', 'Amount', 'Status', 'Processed On'])
        for p in payouts:
            ws4.append([p.requested_on.strftime('%Y-%m-%d %H:%M:%S'), p.student.user.username if p.student and getattr(p.student, 'user', None) else '', float(p.amount), p.status, p.processed_on.strftime('%Y-%m-%d %H:%M:%S') if p.processed_on else ''])

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        resp = HttpResponse(out.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        resp['Content-Disposition'] = f'attachment; filename="financial_report_{ts}.xlsx"'
        return resp
    else:
        # Fallback: CSV of transactions
        si = io.StringIO()
        writer = csv.writer(si)
        writer.writerow(['Created At', 'Title', 'Type', 'Amount', 'Balance After'])
        for t in tx_qs:
            writer.writerow([t.created_at.strftime('%Y-%m-%d %H:%M:%S'), t.title, t.trans_type, f"{t.amount}", f"{t.balance_after or ''}"])
        resp = HttpResponse(si.getvalue().encode('utf-8'), content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = f'attachment; filename="financial_transactions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        return resp


@login_required
def admin2_process_payout(request, payout_id):
    """Approve or reject a student payout request. POST only."""
    user_profile = getattr(request.user, 'profile', None)
    if not is_admin_role(user_profile):
        return JsonResponse({'error': 'Access denied'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    action = request.POST.get('action')  # 'approve' or 'reject'
    try:
        payout = StudentPayout.objects.get(id=int(payout_id))
    except StudentPayout.DoesNotExist:
        return JsonResponse({'error': 'Payout not found'}, status=404)

    if action == 'approve' and payout.status != 'processed':
        payout.processed_on = timezone.now()
        payout.status = 'processed'
        payout.save()

        # Create financial transaction (debit)
        amt = Decimal(payout.amount)
        last = FinancialTransaction.objects.order_by('-created_at').first()
        last_balance = Decimal(last.balance_after) if last and last.balance_after is not None else Decimal('0.00')
        new_balance = last_balance - amt
        FinancialTransaction.objects.create(title=f"Payout to {payout.student.user.username}", amount=amt, trans_type='debit', balance_after=new_balance)
        return JsonResponse({'success': True})
    elif action == 'reject':
        payout.status = 'rejected'
        payout.save()
        return JsonResponse({'success': True})
    else:
        return JsonResponse({'error': 'Invalid action or already processed'}, status=400)


@login_required
def admin2_notifications(request):
    user_profile = getattr(request.user, 'profile', None)
    if not is_admin_role(user_profile):
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')
    # Show recent notifications for admins and allow sending to selected students
    from django.db.models import Count, Max

    if request.method == 'POST':
        # Support both regular form POST and AJAX POST
        title = request.POST.get('title') or request.POST.get('title')
        message_text = request.POST.get('message') or request.POST.get('message')

        # Students may be submitted as multiple 'students' fields or as a comma-separated string
        student_ids = request.POST.getlist('students') or []
        if not student_ids:
            # try comma-separated
            raw = request.POST.get('students') or ''
            if raw:
                student_ids = [s for s in raw.split(',') if s.strip()]

        if title and message_text and student_ids:
            # Create a Notification entry per recipient (no schema changes / no extra migrations)
            created = 0
            for sid in student_ids:
                try:
                    # students in this project are stored as Profile with role='student'
                    prof = Profile.objects.filter(id=int(sid)).select_related('user').first()
                    if prof and getattr(prof, 'user', None):
                        Notification.objects.create(user=prof.user, title=title, message=message_text)
                        created += 1
                except Exception:
                    continue

            # If AJAX (fetch) request, return JSON
            if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.META.get('CONTENT_TYPE','').startswith('application/json'):
                return JsonResponse({'success': True, 'sent': created})

            messages.success(request, f'Notification sent to {created} recipients')
            return redirect('admin2_notifications')

    # Build grouped recent notifications: group by title+message and count recipients
    grouped = (Notification.objects
               .values('title', 'message')
               .annotate(recipients_count=Count('id'), last_sent=Max('created_at'))
               .order_by('-last_sent')[:50])

    # Load students and teachers (only id, name, email, department) for the selection UI
    students_qs = Profile.objects.filter(role='student').select_related('user').order_by('user__username')
    teachers_qs = Profile.objects.filter(role='teacher').select_related('user').order_by('user__username')

    def profile_to_dict(p):
        return {
            'id': p.id,
            'name': (p.user.get_full_name() or p.user.username),
            'email': (p.user.email or ''),
            'department': (p.department or '')
        }

    students = [profile_to_dict(p) for p in students_qs]
    teachers = [profile_to_dict(p) for p in teachers_qs]

    # Normalize grouped entries into a list of dicts
    notifications = []
    for g in grouped:
        notifications.append({'title': g['title'], 'message': g['message'], 'recipients_count': g['recipients_count'], 'created_at': g['last_sent']})

    return render(request, 'dashboards/admin2_notifications.html', {'notifications': notifications, 'students': students, 'teachers': teachers})


# --- Superadmin advanced features: PDF and Excel exports -----------------
@login_required
def export_selected_students_pdf(request):
    """Export selected students to PDF using dynamic imports."""
    import importlib
    
    user_profile = getattr(request.user, 'profile', None)
    if not is_admin_role(user_profile):
        return JsonResponse({'error': 'Access denied'}, status=403)

    ids = request.GET.get('ids', '').split(',')
    if not ids:
        messages.error(request, 'No students selected')
        return redirect('superadmin_full_lists')

    try:
        # Dynamically import reportlab components
        modules = {
            'colors': importlib.import_module('reportlab.lib.colors'),
            'pagesizes': importlib.import_module('reportlab.lib.pagesizes'),
            'platypus': importlib.import_module('reportlab.platypus'),
            'styles_mod': importlib.import_module('reportlab.lib.styles'),
            'units': importlib.import_module('reportlab.lib.units'),
            'enums': importlib.import_module('reportlab.lib.enums')
        }
    except ImportError as e:
        messages.error(request, f'PDF generation requires reportlab. Error: {e}')
        return redirect('superadmin_full_lists')

    # Query selected students
    students = Profile.objects.filter(id__in=ids, role='student').select_related('user')
    
    # Set up document components
    colors = modules['colors']
    A4 = modules['pagesizes'].A4
    landscape = modules['pagesizes'].landscape
    SimpleDocTemplate = modules['platypus'].SimpleDocTemplate
    Paragraph = modules['platypus'].Paragraph
    Spacer = modules['platypus'].Spacer
    Table = modules['platypus'].Table
    TableStyle = modules['platypus'].TableStyle
    ParagraphStyle = modules['styles_mod'].ParagraphStyle
    inch = modules['units'].inch
    getSampleStyleSheet = modules['styles_mod'].getSampleStyleSheet
    TA_CENTER = modules['enums'].TA_CENTER
    
    # Generate PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = []

    # Title with enhanced styling
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1a56db')
    )
    elements.append(Paragraph('Selected Students Report', title_style))
    elements.append(Spacer(1, 0.3*inch))

    # Table data
    data = [['ID', 'Username', 'Full Name', 'Email', 'Department', 'Phone', 'Joined Date']]
    for student in students:
        data.append([
            str(student.user.id),
            student.user.username,
            student.user.get_full_name(),
            student.user.email,
            student.department or '',
            student.phone or '',
            student.user.date_joined.strftime('%Y-%m-%d')
        ])

    # Table data and styling will be handled together

    # Create table with combined styling
    table = Table(data, repeatRows=1)
    table_styles = [
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a5fb4')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 12),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.white),
        ('TEXTCOLOR', (0,1), (-1,-1), colors.black),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,1), (-1,-1), 10),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        # Enhanced styling
        ('BACKGROUND', (0,1), (-1,-1), colors.Color(0.95, 0.95, 0.95)),
        ('TEXTCOLOR', (0,1), (-1,-1), colors.Color(0.2, 0.2, 0.2)),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LEADING', (0,0), (-1,-1), 12)
    ]
    table.setStyle(TableStyle(table_styles))
    elements.append(table)
    
    # Add footer with timestamp
    elements.append(Spacer(1, 0.5*inch))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.gray,
        alignment=TA_CENTER
    )
    footer_text = f'Generated on {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}'
    elements.append(Paragraph(footer_text, footer_style))

    # Build PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()

    # Generate response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="students_report_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf"'
    response.write(pdf)
    return response


@login_required
def export_selected_students_excel(request):
    """Export selected students to Excel."""
    user_profile = getattr(request.user, 'profile', None)
    if not is_admin_role(user_profile):
        return JsonResponse({'error': 'Access denied'}, status=403)

    ids = request.GET.get('ids', '').split(',')
    if not ids:
        messages.error(request, 'No students selected')
        return redirect('superadmin_full_lists')

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        messages.error(request, 'Excel export requires openpyxl')
        return redirect('superadmin_full_lists')

    # Query selected students
    students = Profile.objects.filter(id__in=ids, role='student').select_related('user')

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Students Report"

    # Headers
    headers = ['ID', 'Username', 'Full Name', 'Email', 'Department', 'Phone', 'Joined Date']
    ws.append(headers)

    # Style header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a5fb4", end_color="1a5fb4", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Add data
    for student in students:
        ws.append([
            student.user.id,
            student.user.username,
            student.user.get_full_name(),
            student.user.email,
            student.department or '',
            student.phone or '',
            student.user.date_joined.strftime('%Y-%m-%d')
        ])

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Generate Excel file
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="students_report_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    response.write(buffer.getvalue())
    return response


# --- Superadmin advanced features scaffold (placeholders) -----------------
@login_required
def superadmin_student_management(request):
    """Enhanced Student Management module with search, filter and pagination."""
    user_profile = getattr(request.user, 'profile', None)
    if not (user_profile and getattr(user_profile, 'role', None) in ('superadmin', 'admin2')):
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    # Get base queryset
    students_qs = Profile.objects.filter(role='student').select_related('user').order_by('-user__date_joined')
    
    # Search functionality
    search_query = request.GET.get('search', '').strip()
    department_filter = request.GET.get('department', '').strip()
    # accept new param name 'student_class' but fall back to legacy 'class' for compatibility
    class_filter = request.GET.get('student_class', request.GET.get('class', '')).strip()
    
    if search_query:
        students_qs = students_qs.filter(
            Q(user__username__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(roll_number__icontains=search_query)
        )
    
    if department_filter:
        students_qs = students_qs.filter(department=department_filter)
        
    if class_filter:
        students_qs = students_qs.filter(student_class=class_filter)
    
    # Get unique departments and classes for filters
    departments = list(Profile.objects.filter(role='student')
                      .exclude(department='')
                      .values_list('department', flat=True)
                      .distinct())
    classes = list(Profile.objects.filter(role='student')
                  .exclude(student_class='')
                  .values_list('student_class', flat=True)
                  .distinct())
    
    # Calculate stats
    last_week = timezone.now() - timedelta(days=7)
    total_students = students_qs.count()
    total_departments = len(departments)
    total_classes = len(classes)
    new_students = students_qs.filter(user__date_joined__gte=last_week).count()
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(students_qs, 10)  # Show 10 students per page
    
    try:
        students = paginator.page(page)
    except PageNotAnInteger:
        students = paginator.page(1)
    except EmptyPage:
        students = paginator.page(paginator.num_pages)

    context = {
        'students': students,
        'departments': departments,
        'classes': classes,
        'total_students': total_students,
        'total_departments': total_departments,
        'total_classes': total_classes,
        'new_students': new_students,
        'search_query': search_query,
        'department_filter': department_filter,
        'class_filter': class_filter
    }
    
    return render(request, 'dashboards/superadmin_student_management.html', context)


@login_required
def superadmin_student_add(request):
    user_profile = getattr(request.user, 'profile', None)
    if not (user_profile and getattr(user_profile, 'role', None) == 'superadmin'):
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    from .forms import StudentCreateForm
    if request.method == 'POST':
        form = StudentCreateForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Student created')
            return redirect('superadmin_student_management')
    else:
        form = StudentCreateForm()

    return render(request, 'dashboards/superadmin_student_add.html', {'form': form})


@login_required
def superadmin_student_edit(request, profile_id):
    user_profile = getattr(request.user, 'profile', None)
    if not (user_profile and getattr(user_profile, 'role', None) == 'superadmin'):
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    try:
        profile = Profile.objects.get(id=profile_id, role='student')
    except Profile.DoesNotExist:
        messages.error(request, 'Student not found')
        return redirect('superadmin_student_management')

    from .forms import StudentCreateForm
    if request.method == 'POST':
        form = StudentCreateForm(request.POST, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, 'Student updated')
            return redirect('superadmin_student_management')
    else:
        # Prefill username/email via instance.user
        initial = {}
        if getattr(profile, 'user', None):
            initial.update({'username': profile.user.username, 'email': profile.user.email, 'first_name': profile.user.first_name, 'last_name': profile.user.last_name})
        form = StudentCreateForm(instance=profile, initial=initial)

    return render(request, 'dashboards/superadmin_student_edit.html', {'form': form, 'profile': profile})


@login_required
def superadmin_student_delete(request, profile_id):
    user_profile = getattr(request.user, 'profile', None)
    if not (user_profile and getattr(user_profile, 'role', None) == 'superadmin'):
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    try:
        profile = Profile.objects.get(id=profile_id, role='student')
    except Profile.DoesNotExist:
        messages.error(request, 'Student not found')
        return redirect('superadmin_student_management')

    # On GET, show confirmation. On POST, perform delete.
    if request.method == 'POST':
        user = profile.user
        profile.delete()
        if user:
            user.delete()
        messages.success(request, 'Student deleted')
        return redirect('superadmin_student_management')

    return render(request, 'dashboards/superadmin_student_delete_confirm.html', {'profile': profile})


@login_required
def superadmin_financial_overview(request):
    """Placeholder for fee control, invoices, balance renderer and charts."""
    user_profile = getattr(request.user, 'profile', None)
    if not (user_profile and getattr(user_profile, 'role', None) == 'superadmin'):
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    # Minimal stub data for charts (months, income, expenses)
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun']
    income = [12000, 15000, 14000, 17000, 16000, 18000]
    expenses = [4000, 3000, 3500, 4200, 3800, 4100]

    context = {
        'months': months,
        'income': income,
        'expenses': expenses,
    }
    return render(request, 'dashboards/superadmin_financial.html', context)


@login_required
def superadmin_attendance_live(request):
    """Placeholder for live attendance dashboard; real-time integrations are future work."""
    user_profile = getattr(request.user, 'profile', None)
    if not (user_profile and getattr(user_profile, 'role', None) == 'superadmin'):
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    # Dummy live entries
    live_checkins = [
        {'student': 'student1', 'time': '08:02', 'method': 'QR'},
        {'student': 'student2', 'time': '08:10', 'method': 'Biometric'},
    ]
    return render(request, 'dashboards/superadmin_attendance_live.html', {'live_checkins': live_checkins})


@login_required
def superadmin_updates(request):
    """Return JSON used by superadmin dashboard for live updates.

    Payload: { total_students, total_teachers, total_courses, other_count, recent_users: [{id, username, email, role, joined}], signup_labels:[], signup_counts:[] }
    """
    user_profile = getattr(request.user, 'profile', None)
    if not (user_profile and getattr(user_profile, 'role', None) in ('superadmin', 'admin2')):
        return JsonResponse({'error': 'Access denied'}, status=403)

    # totals
    total_students = Profile.objects.filter(role='student').count()
    total_teachers = Profile.objects.filter(role='teacher').count()
    total_courses = Course.objects.count()
    total_profiles = Profile.objects.count()
    total_users = User.objects.count()
    other_count = max(0, total_users - (total_students + total_teachers))

    # recent users (5)
    recent_qs = User.objects.order_by('-date_joined')[:8]
    recent = []
    for u in recent_qs:
        recent.append({'id': u.id, 'username': u.username, 'email': u.email or '', 'role': getattr(getattr(u, 'profile', None), 'role', ''), 'joined': u.date_joined.strftime('%Y-%m-%d')})

    # signups last 7 days
    today = timezone.now().date()
    labels = []
    counts = []
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        labels.append(d.strftime('%b %d'))
        counts.append(User.objects.filter(date_joined__date=d).count())

    data = {
        'total_students': total_students,
        'total_teachers': total_teachers,
        'total_courses': total_courses,
        'other_count': other_count,
        'recent_users': recent,
        'signup_labels': labels,
        'signup_counts': counts,
    }

    return JsonResponse(data)


# --- Superadmin Staff Attendance (webcam / face attendance MVP) -----------------
@role_required('superadmin')
def superadmin_staff_attendance(request):
    """Render the staff attendance page where admins can capture webcam photos
    and review recent check-ins. This page is intended for superadmins only.
    """
    user_profile = getattr(request.user, 'profile', None)
    if not (user_profile and getattr(user_profile, 'role', None) == 'superadmin'):
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    staff_members = StaffMember.objects.select_related('profile__user').all()
    return render(request, 'dashboards/superadmin_staff_attendance.html', {'staff_members': staff_members})


@role_required(['superadmin','admin2'])
def superadmin_student_list(request):
    """Return rendered HTML for a list of students (used by dashboard modal)."""
    q = (request.GET.get('q') or '').strip()
    students_qs = Profile.objects.filter(role='student').select_related('user').order_by('-user__date_joined')
    if q:
        # if numeric, allow searching by user id as well
        if q.isdigit():
            students_qs = students_qs.filter(Q(user__id=int(q)) | Q(user__username__icontains=q) | Q(user__first_name__icontains=q) | Q(user__last_name__icontains=q) | Q(department__icontains=q))
        else:
            students_qs = students_qs.filter(Q(user__username__icontains=q) | Q(user__first_name__icontains=q) | Q(user__last_name__icontains=q) | Q(department__icontains=q))
    students = students_qs[:500]
    html = render_to_string('dashboards/partials/superadmin_students_list.html', {'students': students}, request=request)
    return JsonResponse({'success': True, 'html': html})


@login_required
def debug_whoami(request):
    """Temporary debug endpoint returning current user/profile info. Remove in production."""
    user = request.user
    profile = getattr(user, 'profile', None)
    data = {
        'id': getattr(user, 'id', None),
        'username': getattr(user, 'username', None),
        'is_authenticated': user.is_authenticated,
        'is_superuser': user.is_superuser,
        'profile_exists': bool(profile),
        'profile_role': getattr(profile, 'role', None) if profile else None,
        'profile_department': getattr(profile, 'department', None) if profile else None,
    }
    return JsonResponse({'success': True, 'whoami': data})


@role_required(['superadmin','admin2'])
def superadmin_full_lists(request):
    """Render a full page with tabs for Students, Staff and Teachers, with filters and export.

    Query params:
      - tab: students|staff|teachers (default students)
      - q: search string (name, username, id, department)
      - department: filter by department
      - page: pagination page number
      - format: csv to download CSV
    """
    tab = (request.GET.get('tab') or 'students').lower()
    q = (request.GET.get('q') or '').strip()
    dept = (request.GET.get('department') or '').strip()
    page = request.GET.get('page')
    fmt = request.GET.get('format')

    students_qs = Profile.objects.filter(role='student').select_related('user').order_by('-user__date_joined')
    staff_qs = StaffMember.objects.select_related('profile__user').order_by('-profile__user__date_joined')
    teachers_qs = Profile.objects.filter(role='teacher').select_related('user').order_by('-user__date_joined')

    if q:
        if q.isdigit():
            students_qs = students_qs.filter(Q(user__id=int(q)) | Q(user__username__icontains=q) | Q(user__first_name__icontains=q) | Q(user__last_name__icontains=q) | Q(department__icontains=q))
            staff_qs = staff_qs.filter(Q(id=int(q)) | Q(profile__user__username__icontains=q) | Q(profile__user__first_name__icontains=q) | Q(profile__user__last_name__icontains=q) | Q(profile__department__icontains=q))
            teachers_qs = teachers_qs.filter(Q(user__id=int(q)) | Q(user__username__icontains=q) | Q(user__first_name__icontains=q) | Q(user__last_name__icontains=q) | Q(department__icontains=q))
        else:
            students_qs = students_qs.filter(Q(user__username__icontains=q) | Q(user__first_name__icontains=q) | Q(user__last_name__icontains=q) | Q(department__icontains=q))
            staff_qs = staff_qs.filter(Q(profile__user__username__icontains=q) | Q(profile__user__first_name__icontains=q) | Q(profile__user__last_name__icontains=q) | Q(profile__department__icontains=q))
            teachers_qs = teachers_qs.filter(Q(user__username__icontains=q) | Q(user__first_name__icontains=q) | Q(user__last_name__icontains=q) | Q(department__icontains=q))

    if dept:
        students_qs = students_qs.filter(department__icontains=dept)
        staff_qs = staff_qs.filter(profile__department__icontains=dept)
        teachers_qs = teachers_qs.filter(department__icontains=dept)

    # Export CSV if requested
    if fmt == 'csv':
        import csv, io
        si = io.StringIO()
        writer = csv.writer(si)
        if tab == 'staff':
            writer.writerow(['Staff ID', 'Username', 'Full name', 'Department', 'Position', 'Joined'])
            for s in staff_qs:
                writer.writerow([s.id, s.profile.user.username if getattr(s, 'profile', None) and getattr(s.profile, 'user', None) else '', (s.profile.user.get_full_name() if getattr(s, 'profile', None) and getattr(s.profile, 'user', None) else ''), s.profile.department if getattr(s, 'profile', None) else '', s.position or '', getattr(getattr(s, 'profile', None) and getattr(s.profile, 'user', None), 'date_joined', '')])
            name = 'staff_list.csv'
        elif tab == 'teachers':
            writer.writerow(['User ID', 'Username', 'Full name', 'Email', 'Department', 'Joined'])
            for t in teachers_qs:
                writer.writerow([t.user.id, t.user.username, t.user.get_full_name(), t.user.email, t.department or '', t.user.date_joined.strftime('%Y-%m-%d') if getattr(t.user, 'date_joined', None) else ''])
            name = 'teachers_list.csv'
        else:
            writer.writerow(['User ID', 'Username', 'Full name', 'Email', 'Department', 'Joined'])
            for u in students_qs:
                writer.writerow([u.user.id, u.user.username, u.user.get_full_name(), u.user.email, u.department or '', u.user.date_joined.strftime('%Y-%m-%d') if getattr(u.user, 'date_joined', None) else ''])
            name = 'students_list.csv'

        resp = HttpResponse(si.getvalue().encode('utf-8'), content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = f'attachment; filename="{name}"'
        return resp

    # Choose dataset for display
    if tab == 'staff':
        dataset = staff_qs
    elif tab == 'teachers':
        dataset = teachers_qs
    else:
        tab = 'students'
        dataset = students_qs

    paginator = Paginator(dataset, 50)
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    context = {
        'tab': tab,
        'q': q,
        'department': dept,
        'page_obj': page_obj,
        'paginator': paginator,
    }
    return render(request, 'dashboards/superadmin_full_lists.html', context)


@role_required(['superadmin','admin2'])
def superadmin_staff_list(request):
    """Return rendered HTML for a list of staff members (used by dashboard modal)."""
    q = (request.GET.get('q') or '').strip()
    staff_qs = StaffMember.objects.select_related('profile__user').order_by('-profile__user__date_joined')
    if q:
        if q.isdigit():
            staff_qs = staff_qs.filter(Q(id=int(q)) | Q(profile__user__username__icontains=q) | Q(profile__user__first_name__icontains=q) | Q(profile__user__last_name__icontains=q) | Q(profile__department__icontains=q))
        else:
            staff_qs = staff_qs.filter(Q(profile__user__username__icontains=q) | Q(profile__user__first_name__icontains=q) | Q(profile__user__last_name__icontains=q) | Q(profile__department__icontains=q))
    staff_qs = staff_qs[:500]
    html = render_to_string('dashboards/partials/superadmin_staff_list.html', {'staff': staff_qs}, request=request)
    return JsonResponse({'success': True, 'html': html})


@role_required(['superadmin','admin2'])
@require_http_methods(['POST'])
def one_click_attendance(request):
    """Mark today's attendance for all StaffMember:
    - If a StaffAttendance record exists today for the staff -> present
    - Otherwise -> absent
    Returns JSON summary and triggers Notifications for absentees.
    """
    today_date = timezone.now().date()
    staff_qs = StaffMember.objects.select_related('profile__user')

    # attendance records captured by webcam/recognition for today
    sa_qs = StaffAttendance.objects.filter(timestamp__date=today_date)

    # Map staff_id -> present
    present_staff_ids = set(sa.staff.id for sa in sa_qs if sa.staff and sa.staff.id)

    created = 0
    updated = 0
    present_count = 0
    absent_count = 0
    leave_count = 0

    for staff in staff_qs:
        status = 'present' if staff.id in present_staff_ids else 'absent'
        SDA, was_created = StaffDailyAttendance.objects.update_or_create(
            staff=staff,
            date=today_date,
            defaults={'status': status, 'method': 'one-click', 'recorded_by': request.user}
        )
        if was_created:
            created += 1
        else:
            updated += 1

        if status == 'present':
            present_count += 1
        elif status == 'absent':
            absent_count += 1
        else:
            leave_count += 1

    # Notify absentees (in-app notification and optional email)
    absentees = StaffDailyAttendance.objects.filter(date=today_date, status='absent').select_related('staff__profile__user')
    notif_title = f"Attendance marked for {today_date.strftime('%Y-%m-%d')}"
    for rec in absentees:
        user = getattr(getattr(rec, 'staff', None), 'profile', None) and getattr(rec.staff.profile, 'user', None)
        if user:
            message = f"You are marked Absent on {today_date.strftime('%Y-%m-%d')}. If this is incorrect, please contact admin."
            Notification.objects.create(user=user, title=notif_title, message=message)
            # optional email (best-effort)
            try:
                if getattr(settings, 'EMAIL_HOST', None):
                    send_mail(subject=notif_title, message=message, from_email=settings.DEFAULT_FROM_EMAIL, recipient_list=[user.email], fail_silently=True)
            except Exception:
                pass

    data = {'created': created, 'updated': updated, 'present': present_count, 'absent': absent_count, 'leave': leave_count}
    return JsonResponse({'success': True, 'summary': data})


@role_required(['superadmin','admin2'])
def admin2_attendance_summary(request):
    """Return JSON summary for attendance (counts for date). Query param 'date' optional (YYYY-MM-DD)."""
    user_profile = getattr(request.user, 'profile', None)
    if not is_admin_role(user_profile):
        return JsonResponse({'error': 'Access denied'}, status=403)

    date_str = request.GET.get('date')
    if date_str:
        try:
            qdate = date.fromisoformat(date_str)
        except Exception:
            return JsonResponse({'error': 'Invalid date'}, status=400)
    else:
        qdate = timezone.now().date()

    total_staff = StaffMember.objects.count()
    present = StaffDailyAttendance.objects.filter(date=qdate, status='present').count()
    absent = StaffDailyAttendance.objects.filter(date=qdate, status='absent').count()
    leave = StaffDailyAttendance.objects.filter(date=qdate, status='leave').count()

    return JsonResponse({'date': qdate.isoformat(), 'total_staff': total_staff, 'present': present, 'absent': absent, 'leave': leave})


@role_required(['superadmin','admin2'])
def admin2_export_attendance(request):
    """Export attendance records filtered by date range, department, or staff id.
    Query params: start, end (YYYY-MM-DD), department, staff_id, format=csv|xlsx
    """
    user_profile = getattr(request.user, 'profile', None)
    if not is_admin_role(user_profile):
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    start_str = request.GET.get('start')
    end_str = request.GET.get('end')
    start_date, end_date, errs = validate_date_range(start_str, end_str)
    if errs:
        messages.warning(request, ' '.join(errs))

    qs = StaffDailyAttendance.objects.select_related('staff__profile__user')
    if start_date:
        qs = qs.filter(date__gte=start_date)
    if end_date:
        qs = qs.filter(date__lte=end_date)

    dept = request.GET.get('department')
    if dept:
        qs = qs.filter(staff__profile__department=dept)

    staff_id = request.GET.get('staff_id')
    if staff_id:
        try:
            qs = qs.filter(staff__id=int(staff_id))
        except Exception:
            pass

    fmt = request.GET.get('format') or 'csv'
    if fmt == 'xlsx':
        try:
            from openpyxl import Workbook  # type: ignore
            wb = Workbook()
            ws = wb.active
            ws.append(['Date', 'Staff', 'Username', 'Department', 'Status', 'Recorded At', 'Method', 'Note'])
            for r in qs.order_by('date'):
                ws.append([r.date.isoformat(), r.staff.profile.user.get_full_name() or r.staff.profile.user.username, r.staff.profile.user.username, r.staff.profile.department or '', r.status, r.timestamp.strftime('%Y-%m-%d %H:%M:%S'), r.method, r.note or ''])
            out = BytesIO()
            wb.save(out)
            out.seek(0)
            resp = HttpResponse(out.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp['Content-Disposition'] = f'attachment; filename="attendance_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            return resp
        except Exception:
            # fallback to csv
            fmt = 'csv'

    # CSV fallback
    si = io.StringIO()
    writer = csv.writer(si)
    writer.writerow(['Date', 'Staff', 'Username', 'Department', 'Status', 'Recorded At', 'Method', 'Note'])
    for r in qs.order_by('date'):
        writer.writerow([r.date.isoformat(), r.staff.profile.user.get_full_name() or r.staff.profile.user.username, r.staff.profile.user.username, r.staff.profile.department or '', r.status, r.timestamp.strftime('%Y-%m-%d %H:%M:%S'), r.method, r.note or ''])

    resp = HttpResponse(si.getvalue().encode('utf-8'), content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="attendance_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    return resp


@role_required(['superadmin','admin2'])
def admin2_attendance_ai_insights(request):
    """Basic heuristic AI insights for attendance: top punctual staff and 30-day absentee trend."""
    user_profile = getattr(request.user, 'profile', None)
    if not is_admin_role(user_profile):
        return JsonResponse({'error': 'Access denied'}, status=403)

    # last 30 days
    end = timezone.now().date()
    start = end - timedelta(days=29)

    # absentee trend per day
    trend = []
    dates = []
    for i in range(0, 30):
        d = start + timedelta(days=i)
        dates.append(d.isoformat())
        absent = StaffDailyAttendance.objects.filter(date=d, status='absent').count()
        trend.append(absent)

    # top punctual staff: use StaffAttendance earliest timestamp per day avg
    punctual_scores = {}
    sa_qs = StaffAttendance.objects.filter(timestamp__date__gte=start, timestamp__date__lte=end).select_related('staff__profile__user')
    # compute first-checkin per staff per day
    first_per_staff_day = {}
    for sa in sa_qs:
        if not sa.staff:
            continue
        key = (sa.staff.id, sa.timestamp.date())
        if key not in first_per_staff_day or sa.timestamp < first_per_staff_day[key]:
            first_per_staff_day[key] = sa.timestamp

    # aggregate average epoch per staff
    import statistics
    by_staff = {}
    for (staff_id, d), ts in first_per_staff_day.items():
        by_staff.setdefault(staff_id, []).append(ts.hour * 3600 + ts.minute * 60 + ts.second)

    punctual_list = []
    for staff_id, secs in by_staff.items():
        try:
            avg_secs = statistics.mean(secs)
            staff = StaffMember.objects.filter(id=staff_id).select_related('profile__user').first()
            punctual_list.append({'staff_id': staff_id, 'username': staff.profile.user.username if staff else '', 'avg_seconds': avg_secs})
        except Exception:
            continue

    punctual_list.sort(key=lambda x: x['avg_seconds'])
    top5 = punctual_list[:5]

    return JsonResponse({'trend_dates': dates, 'absent_trend': trend, 'top_punctual': top5})


@role_required(['superadmin','admin2'])
def admin2_attendance_list(request):
    """Return JSON list of staff members with today's attendance status and recent history.

    Query params:
      - date (YYYY-MM-DD) optional (defaults to today)
      - department optional
      - staff_id optional
      - days optional: number of days of history (default 7)
    """
    date_str = request.GET.get('date')
    if date_str:
        try:
            qdate = date.fromisoformat(date_str)
        except Exception:
            return JsonResponse({'error': 'Invalid date'}, status=400)
    else:
        qdate = timezone.now().date()

    days = int(request.GET.get('days') or 7)
    dept = request.GET.get('department')
    staff_filter = request.GET.get('staff_id')

    staffs = StaffMember.objects.select_related('profile__user')
    if dept:
        staffs = staffs.filter(profile__department=dept)
    if staff_filter:
        try:
            staffs = staffs.filter(id=int(staff_filter))
        except Exception:
            pass

    # build date window
    dates = [(qdate - timedelta(days=i)) for i in range(days-1, -1, -1)]

    out = []
    for s in staffs:
        user = getattr(s.profile, 'user', None)
        username = user.username if user else ''
        full = (user.get_full_name() if user else '')
        photo = None
        try:
            if getattr(s.profile, 'profile_pic', None):
                photo = request.build_absolute_uri(s.profile.profile_pic.url)
        except Exception:
            photo = None

        # today's status
        today_rec = StaffDailyAttendance.objects.filter(staff=s, date=qdate).first()
        status = today_rec.status if today_rec else 'unknown'
        last_att = StaffDailyAttendance.objects.filter(staff=s).order_by('-date').first()
        last_ts = last_att.timestamp.isoformat() if last_att else None

        # recent history
        history_qs = StaffDailyAttendance.objects.filter(staff=s, date__in=dates)
        history_map = {h.date.isoformat(): h.status for h in history_qs}
        history = [history_map.get(d.isoformat(), '') for d in dates]

        out.append({'staff_id': s.id, 'username': username, 'full_name': full, 'department': s.profile.department or '', 'position': s.position or '', 'photo_url': photo, 'status': status, 'last_attendance': last_ts, 'history_dates': [d.isoformat() for d in dates], 'history': history})

    return JsonResponse({'date': qdate.isoformat(), 'days': days, 'records': out})


@role_required(['superadmin','admin2','staff'])
@require_http_methods(["POST"])
def superadmin_staff_attendance_recognize(request):
    """Accept an uploaded image (multipart/form-data 'image') or an optional
    'recognized_username' field and create a StaffAttendance record.

    This endpoint intentionally keeps recognition logic as a simple heuristic
    for the MVP: if a recognized_username is supplied it will be linked to a
    StaffMember when possible. For real face recognition integrate a dedicated
    library or external service and replace this handler.
    """
    # Allow any authenticated user (staff or admin) to POST a checkin image
    image = request.FILES.get('image')
    recognized = request.POST.get('recognized_username') or request.POST.get('username')

    attendance = StaffAttendance()
    attendance.method = 'webcam'
    if recognized:
        attendance.recognized_username = recognized
        # try to attach a StaffMember if the username exists
        try:
            profile = Profile.objects.select_related('user').get(user__username=recognized)
            staff = StaffMember.objects.filter(profile=profile).first()
            if staff:
                attendance.staff = staff
        except Exception:
            pass

    if image:
        attendance.image = image

    attendance.save()

    data = {
        'success': True,
        'id': attendance.id,
        'timestamp': attendance.timestamp.isoformat(),
        'recognized_username': attendance.recognized_username or '',
        'staff_id': attendance.staff.id if attendance.staff else None,
    }
    return JsonResponse(data)


@role_required(['superadmin','admin2'])
def staff_attendance_marked(request):
    """Render the staff attendance marked page."""
    user_profile = getattr(request.user, 'profile', None)
    if not is_admin_role(user_profile):
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')
    return render(request, 'dashboards/staff_attendance_marked.html')


@role_required(['superadmin','admin2'])
def superadmin_staff_attendance_updates(request):
    """Return recent staff attendance records as JSON for dashboard polling.

    Query params:
      - n (optional): number of recent records to return (default 20)
    """
    user_profile = getattr(request.user, 'profile', None)
    if not (user_profile and getattr(user_profile, 'role', None) in ('superadmin', 'admin2')):
        return JsonResponse({'error': 'Access denied'}, status=403)

    try:
        n = int(request.GET.get('n') or 20)
    except Exception:
        n = 20

    records = StaffAttendance.objects.select_related('staff__profile__user').all()[:n]
    out = []
    for r in records:
        user_display = r.recognized_username or (r.staff.profile.user.username if r.staff and getattr(r.staff, 'profile', None) and getattr(r.staff.profile, 'user', None) else 'unknown')
        img_url = None
        if getattr(r, 'image', None):
            try:
                img_url = request.build_absolute_uri(r.image.url)
            except Exception:
                img_url = None
        out.append({'id': r.id, 'timestamp': r.timestamp.isoformat(), 'user': user_display, 'method': r.method, 'image_url': img_url})

    return JsonResponse({'records': out})


# Teacher Dashboard
@role_required('teacher')
def teacher_dashboard(request):
    
    teacher_profile = request.user.profile
    # annotate courses with enrolled student counts for display
    courses = list(Course.objects.filter(teacher=teacher_profile))
    course_stats = []
    for c in courses:
        student_count = Enrollment.objects.filter(course=c).count()
        # attach a helpful attribute so templates can show enrolled_students
        setattr(c, 'enrolled_students', student_count)
        course_stats.append({'course': c, 'student_count': student_count})

    # total unique students taught by this teacher
    total_students = Enrollment.objects.filter(course__teacher=teacher_profile).values('student').distinct().count()

    # classes_today (not modeled) — placeholder 0
    classes_today = 0

    # average attendance across this teacher's courses
    attendance_percentages = []
    for c in courses:
        total = Attendance.objects.filter(course=c).count()
        present = Attendance.objects.filter(course=c, status=True).count()
        if total > 0:
            attendance_percentages.append((present / total) * 100)

    avg_attendance = round(sum(attendance_percentages) / len(attendance_percentages), 2) if attendance_percentages else 0.0
    # Determine top course by enrollment
    top_course = None
    if course_stats:
        sorted_stats = sorted(course_stats, key=lambda x: x['student_count'], reverse=True)
        top_course = sorted_stats[0]['course']

    # Top students by average marks across this teacher's courses
    from django.db.models import Avg
    today = date.today()
    top_students = []
    try:
        avg_per_student = Submission.objects.filter(assignment__course__teacher=teacher_profile, marks_obtained__isnull=False).values('student').annotate(avg_marks=Avg('marks_obtained')).order_by('-avg_marks')[:5]
        for entry in avg_per_student:
            try:
                student_profile = Profile.objects.get(id=entry['student'])
                top_students.append({'student': student_profile, 'avg_marks': round(entry['avg_marks'] or 0.0, 2)})
            except Profile.DoesNotExist:
                continue
    except Exception:
        top_students = []

    # Upcoming assignments for this teacher's courses — include null due_date so drafts appear
    from django.db.models import Q
    upcoming_assignments = Assignment.objects.filter(course__teacher=teacher_profile).filter(Q(due_date__gte=today) | Q(due_date__isnull=True)).order_by('due_date')[:5]

    context = {
        'courses': courses,
        'total_courses': len(courses),
        'course_stats': course_stats,
        'total_students': total_students,
        'classes_today': classes_today,
        'avg_attendance': avg_attendance,
        'top_course': top_course,
        'teacher_photo_url': getattr(teacher_profile, 'profile_pic', None).url if getattr(teacher_profile, 'profile_pic', None) else None,
        'teacher_email': request.user.email,
        'teacher_phone': getattr(teacher_profile, 'phone', ''),
        'top_students': top_students,
        'upcoming_assignments': upcoming_assignments,
    }
    # Add a slug for linking to the instructor profile (slugified full name or username)
    from django.utils.text import slugify
    display_name = f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username
    context['teacher_slug'] = slugify(display_name)
    return render(request, 'dashboards/teacher_dashboard.html', context)


@role_required('teacher')
def teacher_dashboard_updates(request):
    """Return small JSON payload used by the teacher dashboard for live updates.

    Returns: { total_courses, total_students, pending_tasks, avg_attendance, recent_activities: [{time, title, detail}, ...] }
    """
    # Only teachers may access this endpoint
    if getattr(request.user.profile, 'role', None) != 'teacher':
        return JsonResponse({'error': 'Access denied'}, status=403)

    teacher_profile = request.user.profile
    today = date.today()

    # Core counts
    courses_qs = Course.objects.filter(teacher=teacher_profile)
    total_courses = courses_qs.count()
    total_students = Enrollment.objects.filter(course__teacher=teacher_profile).values('student').distinct().count()

    # Pending assignments (due today or later). Include assignments with no due_date as pending.
    try:
        from django.db.models import Q
        pending_tasks = Assignment.objects.filter(course__teacher=teacher_profile).filter(Q(due_date__gte=today) | Q(due_date__isnull=True)).count()
    except Exception:
        pending_tasks = 0

    # Average attendance across teacher's courses
    attendance_percentages = []
    for c in courses_qs:
        total = Attendance.objects.filter(course=c).count()
        present = Attendance.objects.filter(course=c, status=True).count()
        if total > 0:
            attendance_percentages.append((present / total) * 100)

    avg_attendance = round(sum(attendance_percentages) / len(attendance_percentages), 2) if attendance_percentages else 0.0

    # Recent activity: merge recent submissions, attendance entries and assignments
    events = []
    try:
        subs = Submission.objects.filter(assignment__course__teacher=teacher_profile).select_related('student__user', 'assignment').order_by('-submission_date')[:5]
        for s in subs:
            uname = getattr(getattr(s, 'student', None), 'user', None)
            uname = uname.username if uname else ''
            events.append({'time': s.submission_date.isoformat(), 'title': f"{uname} submitted {s.assignment.title}", 'detail': ''})
    except Exception:
        pass

    try:
        atts = Attendance.objects.filter(course__teacher=teacher_profile).select_related('student__user', 'course').order_by('-date')[:5]
        for a in atts:
            uname = getattr(getattr(a, 'student', None), 'user', None)
            uname = uname.username if uname else ''
            events.append({'time': a.date.isoformat(), 'title': f"Attendance recorded for {a.course.code}", 'detail': f"{uname} - {'Present' if a.status else 'Absent'}"})
    except Exception:
        pass

    try:
        assigns = Assignment.objects.filter(course__teacher=teacher_profile).order_by('-created_at')[:5]
        for asg in assigns:
            time_iso = asg.created_at.isoformat() if getattr(asg, 'created_at', None) else today.isoformat()
            events.append({'time': time_iso, 'title': f"Assignment created: {asg.title}", 'detail': f"For {asg.course.code}"})
    except Exception:
        pass

    # Sort and limit
    events_sorted = sorted(events, key=lambda x: x.get('time', ''), reverse=True)[:5]

    data = {
        'total_courses': total_courses,
        'total_students': total_students,
        'pending_tasks': pending_tasks,
        'avg_attendance': avg_attendance,
        'recent_activities': events_sorted,
    }

    return JsonResponse(data)



@login_required
def add_course(request):
    # Only teachers can add courses
    if getattr(request.user.profile, 'role', None) != 'teacher':
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    if request.method == 'POST':
        name = request.POST.get('name')
        code = request.POST.get('code') or name[:6].upper().replace(' ', '')
        credits = int(request.POST.get('credits') or 3)
        semester = int(request.POST.get('semester') or 1)

        # create course
        try:
            Course.objects.create(name=name, code=code, teacher=request.user.profile, credits=credits, semester=semester)
            messages.success(request, 'Course created successfully')
            return redirect('teacher_dashboard')
        except Exception as e:
            messages.error(request, f'Error creating course: {e}')

    return render(request, 'add_course.html')


@login_required
def create_assignment(request):
    """Create a new assignment for a teacher-owned course. Accepts POST (AJAX-friendly).
    Expects: course_id, title, due_date (YYYY-MM-DD), max_marks
    Returns JSON: {'success': True, 'assignment_id': id} or error payload.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    if getattr(request.user.profile, 'role', None) != 'teacher':
        return JsonResponse({'error': 'Access denied'}, status=403)

    course_id = request.POST.get('course_id')
    title = request.POST.get('title', '').strip()
    due_date = request.POST.get('due_date')
    max_marks = request.POST.get('max_marks') or 100

    if not title:
        return JsonResponse({'error': 'Title is required'}, status=400)

    # Determine course
    try:
        if course_id:
            course = Course.objects.get(id=int(course_id), teacher=request.user.profile)
        else:
            # fallback: first course owned by teacher
            course = Course.objects.filter(teacher=request.user.profile).first()
        if not course:
            return JsonResponse({'error': 'Course not found or access denied'}, status=404)
    except Exception:
        return JsonResponse({'error': 'Invalid course'}, status=400)

    # Parse due_date: accept YYYY-MM-DD or datetime-local value like 2025-11-17T08:34
    due_date_obj = None
    if due_date:
        from datetime import datetime, date
        try:
            # datetime.fromisoformat handles both date and datetime strings (with optional offset)
            parsed = None
            try:
                parsed = datetime.fromisoformat(due_date)
            except Exception:
                # fallback to date.fromisoformat for plain dates
                parsed = date.fromisoformat(due_date)

            if isinstance(parsed, datetime):
                due_date_obj = parsed.date()
            elif isinstance(parsed, date):
                due_date_obj = parsed
        except Exception:
            return JsonResponse({'error': 'Invalid due_date format. Use YYYY-MM-DD or a datetime-local value.'}, status=400)

    # Create assignment
    try:
        a = Assignment.objects.create(
            course=course,
            title=title,
            description=request.POST.get('description', ''),
            body=request.POST.get('body', '') or request.POST.get('description', ''),
            due_date=due_date_obj,
            max_marks=int(max_marks),
            is_draft=(request.POST.get('is_draft') == 'true')
        )

        # Handle uploaded files (assignment_files input supports multiple)
        try:
            files = request.FILES.getlist('assignment_files') or []
            for f in files:
                AssignmentAttachment.objects.create(assignment=a, file=f, uploaded_by=request.user.profile)
        except Exception:
            # Continue even if attachments fail; don't block assignment creation
            pass

        # If imported_file_url provided (from import-from-url flow), record as a simple attachment placeholder
        imported = request.POST.get('imported_file_url')
        if imported:
            # Keep a lightweight representation by saving as StudyMaterial-like file via remote fetch is out-of-scope here.
            # Instead, store the URL in the assignment description for teachers/admins to manage later.
            a.description = (a.description or '') + f"\nImported: {imported}"
            a.save()

        # Questions JSON parsing: frontend sends a 'questions' field with JSON array
        questions_json = request.POST.get('questions')
        if questions_json:
            try:
                import json
                from .models import Question, Option

                qlist = json.loads(questions_json)
                for qdata in qlist:
                    try:
                        q_order = int(qdata.get('order') or 0)
                    except Exception:
                        q_order = 0
                    qtype = qdata.get('type') or 'mcq'
                    qtext = qdata.get('text') or ''
                    try:
                        qpoints = float(qdata.get('points') or 0)
                    except Exception:
                        qpoints = 0

                    qobj = Question.objects.create(
                        assignment=a,
                        order=q_order,
                        qtype=qtype,
                        text=qtext,
                        points=qpoints
                    )

                    # Save options for MCQ and True/False
                    if qtype == 'mcq':
                        opts = qdata.get('options') or []
                        try:
                            correct = int(qdata.get('correct_answer')) if qdata.get('correct_answer') is not None else None
                        except Exception:
                            correct = None
                        for idx, opt_text in enumerate(opts):
                            Option.objects.create(
                                question=qobj,
                                order=idx,
                                text=opt_text,
                                is_correct=(correct == idx)
                            )
                    elif qtype == 'true_false':
                        ca = qdata.get('correct_answer')
                        Option.objects.create(question=qobj, order=0, text='True', is_correct=(ca == 'true' or ca is True))
                        Option.objects.create(question=qobj, order=1, text='False', is_correct=(ca == 'false' or ca is False))
            except Exception:
                # non-fatal: do not block assignment creation on question parsing errors
                pass

        return JsonResponse({'success': True, 'assignment_id': a.id})
    except Exception as e:
        return JsonResponse({'error': f'Error creating assignment: {e}'}, status=500)


@login_required
def assignment_detail(request, assignment_id):
    """Show assignment details and submissions to the owning teacher."""
    profile = getattr(request.user, 'profile', None)
    if not profile or getattr(profile, 'role', None) != 'teacher':
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    try:
        assignment = Assignment.objects.select_related('course').get(id=assignment_id)
    except Assignment.DoesNotExist:
        messages.error(request, 'Assignment not found')
        return redirect('teacher_dashboard')

    # Only the teacher who owns the course may view
    if assignment.course.teacher != profile and not is_admin_role(getattr(request.user, 'profile', None)):
        messages.error(request, 'Access denied')
        return redirect('teacher_dashboard')

    submissions = Submission.objects.filter(assignment=assignment).select_related('student__user').order_by('-submission_date')

    # Allow teacher to grade/update marks via POST (simple form)
    if request.method == 'POST':
        # Expecting fields like marks_<submission_id>
        updated = 0
        for sub in submissions:
            key = f'marks_{sub.id}'
            if key in request.POST:
                try:
                    val = request.POST.get(key)
                    if val is None or val == '':
                        sub.marks_obtained = None
                    else:
                        sub.marks_obtained = float(val)
                    sub.save()
                    updated += 1
                except Exception:
                    continue
        messages.success(request, f'Updated {updated} submissions')
        return redirect('assignment_detail', assignment_id=assignment.id)

    # Summary stats
    graded_count = submissions.filter(marks_obtained__isnull=False).count()
    avg_score = submissions.filter(marks_obtained__isnull=False).aggregate(Avg('marks_obtained'))['marks_obtained__avg'] or 0

    context = {
        'assignment': assignment,
        'submissions': submissions,
        'graded_count': graded_count,
        'avg_score': round(avg_score, 2),
    }
    return render(request, 'assignments/assignment_detail.html', context)


@login_required
def download_assignment_report(request, assignment_id):
    """Download CSV report for an assignment's submissions."""
    profile = getattr(request.user, 'profile', None)
    if not profile or getattr(profile, 'role', None) not in ('teacher', 'superadmin', 'admin2'):
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    try:
        assignment = Assignment.objects.select_related('course').get(id=assignment_id)
    except Assignment.DoesNotExist:
        messages.error(request, 'Assignment not found')
        return redirect('teacher_dashboard')

    if assignment.course.teacher != profile and not is_admin_role(getattr(request.user, 'profile', None)):
        messages.error(request, 'Access denied')
        return redirect('teacher_dashboard')

    submissions = Submission.objects.filter(assignment=assignment).select_related('student__user')

    import csv
    import io

    si = io.StringIO()
    writer = csv.writer(si)
    writer.writerow(['Username', 'Full name', 'Email', 'Submitted at', 'Marks Obtained'])
    for s in submissions:
        uname = s.student.user.username if getattr(s.student, 'user', None) else ''
        full = f"{getattr(getattr(s.student, 'user', None),'first_name','') or ''} {getattr(getattr(s.student, 'user', None),'last_name','') or ''}".strip()
        email = getattr(getattr(s.student, 'user', None), 'email', '') or ''
        submitted = s.submission_date.strftime('%Y-%m-%d %H:%M:%S') if getattr(s, 'submission_date', None) else ''
        marks = s.marks_obtained if s.marks_obtained is not None else ''
        writer.writerow([uname, full, email, submitted, marks])

    csv_bytes = si.getvalue().encode('utf-8')
    resp = HttpResponse(csv_bytes, content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="assignment_{assignment.id}_report.csv"'
    return resp


@login_required
def render_create_assignment(request):
    """Render a full page form for creating an assignment (teacher only)."""
    profile = getattr(request.user, 'profile', None)
    if not profile or getattr(profile, 'role', None) != 'teacher':
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    # Provide list of teacher's courses for selection
    courses = Course.objects.filter(teacher=profile)
    return render(request, 'assignments/create_assignment.html', {'courses': courses})


@login_required
@require_http_methods(["POST"])
def import_from_url(request):
    """Accept a POST with 'url' and return JSON {success: True, file_url: ...}.

    This is a lightweight helper used by the create-assignment page to import
    public documents. Currently it does not fetch or store the content; it
    validates the URL and echoes it back as `file_url`. We keep it simple so
    the frontend can proceed and the backend can be extended later.
    """
    profile = getattr(request.user, 'profile', None)
    if not profile or getattr(profile, 'role', None) != 'teacher':
        return JsonResponse({'error': 'Access denied'}, status=403)

    url = request.POST.get('url', '').strip()
    if not url:
        return JsonResponse({'error': 'URL is required'}, status=400)

    # Basic validation
    if not (url.startswith('http://') or url.startswith('https://')):
        return JsonResponse({'error': 'Invalid URL'}, status=400)

    # For now, return the provided URL as the file reference. Future: fetch/save.
    return JsonResponse({'success': True, 'file_url': url})


@login_required
def course_assignments(request, course_id):
    try:
        course = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        messages.error(request, 'Course not found')
        return redirect('teacher_dashboard')

    profile = getattr(request.user, 'profile', None)
    # Only teacher or superadmin can view the assignments page for the course
    if (not is_admin_role(profile)) and course.teacher != profile:
        messages.error(request, 'Access denied')
        return redirect('role_redirect')

    assignments = Assignment.objects.filter(course=course).order_by('-created_at')

    return render(request, 'assignments/course_assignments.html', {'course': course, 'assignments': assignments})


@login_required
def generate_course_report(request, course_id):
    """Generate a simple CSV report for a teacher's course: student list and avg marks."""
    if getattr(request.user.profile, 'role', None) != 'teacher':
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    try:
        course = Course.objects.get(id=course_id, teacher=request.user.profile)
    except Course.DoesNotExist:
        messages.error(request, 'Course not found')
        return redirect('teacher_dashboard')

    # Optional date filters (YYYY-MM-DD) to limit which assignments contribute to averages
    start = request.GET.get('start')
    end = request.GET.get('end')
    date_filter = {}
    if start:
        try:
            start_date = date.fromisoformat(start)
        except Exception:
            start_date = None
    else:
        start_date = None
    if end:
        try:
            end_date = date.fromisoformat(end)
        except Exception:
            end_date = None
    else:
        end_date = None

    # Build CSV in-memory using a text buffer so encoding is correct.
    import csv
    import io

    si = io.StringIO()
    writer = csv.writer(si)
    # Include contact fields and photo URL so teachers have full details
    writer.writerow(['Username', 'Full name', 'Email', 'Phone', 'Department', 'Photo', 'Avg Marks'])

    enrollments = Enrollment.objects.filter(course=course).select_related('student__user')
    for e in enrollments:
        student = e.student  # Profile instance
        # Average marks for this student in this course (optionally filtered by assignment due_date range)
        subs_qs = Submission.objects.filter(student=student, assignment__course=course, marks_obtained__isnull=False)
        if start_date or end_date:
            assigns = Assignment.objects.filter(course=course)
            if start_date:
                assigns = assigns.filter(due_date__gte=start_date)
            if end_date:
                assigns = assigns.filter(due_date__lte=end_date)
            subs_qs = subs_qs.filter(assignment__in=assigns)
        avg_marks = subs_qs.aggregate(Avg('marks_obtained'))['marks_obtained__avg']
        avg_marks = round(avg_marks or 0.0, 2)

        # Contact fields and photo
        email = (student.user.email or '') if getattr(student, 'user', None) else ''
        phone = getattr(student, 'phone', '') or ''
        department = getattr(student, 'department', '') or ''
        photo_url = ''
        if getattr(student, 'profile_pic', None):
            try:
                raw = student.profile_pic.url
                # Build absolute URL so the CSV is usable outside the site
                photo_url = request.build_absolute_uri(raw)
            except Exception:
                photo_url = ''

        username = (student.user.username if getattr(student, 'user', None) else '')
        full_name = (f"{student.user.first_name} {student.user.last_name}".strip() if getattr(student, 'user', None) else '')

        writer.writerow([username, full_name, email, phone, department, photo_url, avg_marks])

    csv_bytes = si.getvalue().encode('utf-8')
    resp = HttpResponse(csv_bytes, content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="course_{course.code}_report.csv"'
    return resp



@login_required
@require_http_methods(["GET"])
def course_students_json(request, course_id):
    """
    Return JSON list of enrolled students for a course.
    """
    course, error = check_course_access(request.user, course_id)
    if error:
        status_code = 404 if 'not found' in error else 403
        return JsonResponse({'error': error}, status=status_code)

    cache_key = f'course_{course_id}_students_json'
    cached_data = cache.get(cache_key)
    if cached_data:
        return JsonResponse(cached_data)

    enrollments = (Enrollment.objects
                   .filter(course=course)
                   .select_related('student__user')
                   .order_by('student__user__last_name'))

    students = []
    for enrollment in enrollments:
        student = enrollment.student
        user = getattr(student, 'user', None)
        if not user:
            continue
        avg_marks = (Submission.objects
                     .filter(student=student, assignment__course=course, marks_obtained__isnull=False)
                     .aggregate(avg=Avg('marks_obtained'))['avg'])
        students.append({
            'id': student.id,
            'username': user.username,
            'full_name': user.get_full_name() or user.username,
            'email': user.email or '',
            'avg_marks': round(avg_marks or 0.0, 2)
        })

    response_data = {'students': students, 'total': len(students)}
    cache.set(cache_key, response_data, 300)
    return JsonResponse(response_data)


@login_required
@require_http_methods(["GET"])
def teacher_student_report(request, course_id, student_id):
    """
    Generate detailed CSV report for a single student's performance.
    """
    course, error = check_course_access(request.user, course_id)
    if error:
        messages.error(request, error)
        return redirect('teacher_dashboard' if 'not found' in error else 'role_redirect')

    try:
        student = Profile.objects.select_related('user').get(id=student_id, role='student')
    except Profile.DoesNotExist:
        messages.error(request, 'Student not found')
        return redirect('course_assignments', course_id=course.id)

    if not Enrollment.objects.filter(course=course, student=student).exists():
        messages.error(request, 'Student not enrolled in this course')
        return redirect('course_assignments', course_id=course.id)

    start_str = request.GET.get('start')
    end_str = request.GET.get('end')
    start_date, end_date, errors = validate_date_range(start_str, end_str)
    if errors:
        messages.warning(request, ' '.join(errors))

    submissions = (Submission.objects
                   .filter(student=student, assignment__course=course)
                   .select_related('assignment')
                   .order_by('-submission_date'))
    if start_date:
        submissions = submissions.filter(submission_date__date__gte=start_date)
    if end_date:
        submissions = submissions.filter(submission_date__date__lte=end_date)

    si = io.StringIO()
    writer = csv.writer(si)

    user = student.user
    writer.writerow(['Student Report'])
    writer.writerow(['Name:', user.get_full_name() if user else 'N/A'])
    writer.writerow(['Username:', user.username if user else 'N/A'])
    writer.writerow(['Course:', f'{course.name} ({course.code})'])
    writer.writerow(['Report Date:', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
    writer.writerow([])

    writer.writerow(['Assignment ID', 'Assignment Title', 'Due Date', 'Submitted At', 'Days Late', 'Marks Obtained', 'Max Marks', 'Percentage', 'Feedback'])
    total_marks = 0
    total_max = 0
    for sub in submissions:
        assignment = sub.assignment
        due_date = getattr(assignment, 'due_date', None)
        max_marks = getattr(assignment, 'max_marks', 100)
        days_late = ''
        if due_date and sub.submission_date:
            delta = (sub.submission_date.date() - due_date).days
            days_late = max(0, delta)
        percentage = ''
        if sub.marks_obtained is not None and max_marks:
            percentage = round((sub.marks_obtained / max_marks) * 100, 2)
            total_marks += sub.marks_obtained
            total_max += max_marks
        writer.writerow([
            assignment.id,
            assignment.title,
            due_date.strftime('%Y-%m-%d') if due_date else 'N/A',
            sub.submission_date.strftime('%Y-%m-%d %H:%M:%S') if sub.submission_date else 'Not submitted',
            days_late,
            sub.marks_obtained if sub.marks_obtained is not None else 'Not graded',
            max_marks,
            f'{percentage}%' if percentage else 'N/A',
            sub.feedback or ''
        ])

    writer.writerow([])
    writer.writerow(['Summary Statistics'])
    writer.writerow(['Total Submissions:', submissions.count()])
    writer.writerow(['Graded Submissions:', submissions.filter(marks_obtained__isnull=False).count()])
    if total_max > 0:
        overall_pct = round((total_marks / total_max) * 100, 2)
        writer.writerow(['Overall Percentage:', f'{overall_pct}%'])

    log_report_generation(request.user, 'student_csv', course_id, {'student_id': student_id, 'start': start_str, 'end': end_str})

    csv_bytes = si.getvalue().encode('utf-8-sig')
    response = HttpResponse(csv_bytes, content_type='text/csv; charset=utf-8-sig')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'student_{user.username if user else student_id}_course_{course.code}_{timestamp}.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def generate_assignment_averages(request, course_id):
    """Generate CSV with per-assignment average marks for a course.

    Optional GET params: start, end (YYYY-MM-DD) to filter assignments by due_date.
    """
    profile = getattr(request.user, 'profile', None)
    try:
        course = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        messages.error(request, 'Course not found')
        return redirect('teacher_dashboard')

    if (not is_admin_role(profile)) and course.teacher != profile:
        messages.error(request, 'Access denied')
        return redirect('role_redirect')

    start = request.GET.get('start')
    end = request.GET.get('end')
    assigns = Assignment.objects.filter(course=course).order_by('due_date')
    if start:
        try:
            sdate = date.fromisoformat(start)
            assigns = assigns.filter(due_date__gte=sdate)
        except Exception:
            pass
    if end:
        try:
            edate = date.fromisoformat(end)
            assigns = assigns.filter(due_date__lte=edate)
        except Exception:
            pass

    import csv
    import io

    si = io.StringIO()
    writer = csv.writer(si)
    writer.writerow(['Assignment ID', 'Title', 'Due Date', 'Avg Marks', 'Submission Count'])
    for a in assigns:
        stats = Submission.objects.filter(assignment=a, marks_obtained__isnull=False).aggregate(avg=Avg('marks_obtained'), cnt=Count('id'))
        avg = round(stats['avg'] or 0.0, 2)
        cnt = stats['cnt'] or 0
        writer.writerow([a.id, a.title, a.due_date, avg, cnt])

    csv_bytes = si.getvalue().encode('utf-8')
    resp = HttpResponse(csv_bytes, content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="course_{course.code}_assignment_averages.csv"'
    return resp


@login_required
def generate_course_report_pdf(request, course_id):
    """Generate a PDF summary for a course: students, contact, avg marks, joined date, attendance."""
    if getattr(request.user.profile, 'role', None) != 'teacher':
        from django.contrib import messages
        messages.error(request, 'Access denied!')
        from django.shortcuts import redirect
        return redirect('role_redirect')

    try:
        course = Course.objects.get(id=course_id, teacher=request.user.profile)
    except Course.DoesNotExist:
        from django.contrib import messages
        messages.error(request, 'Course not found')
        from django.shortcuts import redirect
        return redirect('teacher_dashboard')

    # Lazy import reportlab to avoid import-time errors
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except Exception:
        from django.contrib import messages
        messages.error(request, 'reportlab is required to generate PDF reports')
        return redirect('teacher_dashboard')

    from .models import Enrollment, Submission, Attendance

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f'Course Report: {course.name} ({course.code})', styles['Heading1']))
    story.append(Spacer(1, 0.2*inch))

    # Table header
    data = [['Username', 'Full name', 'Email', 'Phone', 'Joined', 'Avg Marks', 'Attendance %']]

    enrollments = Enrollment.objects.filter(course=course).select_related('student__user')
    for e in enrollments:
        student = e.student
        user = getattr(student, 'user', None)
        username = user.username if user else ''
        full = f"{getattr(user,'first_name','') or ''} {getattr(user,'last_name','') or ''}".strip() if user else ''
        email = user.email if user else ''
        phone = getattr(student, 'phone', '') or ''
        joined = ''
        try:
            joined = user.date_joined.strftime('%Y-%m-%d') if user else ''
        except Exception:
            joined = ''
        avg_marks = Submission.objects.filter(student=student, assignment__course=course, marks_obtained__isnull=False).aggregate(Avg('marks_obtained'))['marks_obtained__avg']
        avg_marks = (round(avg_marks,2) if avg_marks else 0.0)

        # Attendance percentage for this student in this course
        total = Attendance.objects.filter(student=student, course=course).count()
        present = Attendance.objects.filter(student=student, course=course, status=True).count()
        attendance_pct = round((present / total * 100), 2) if total > 0 else 0.0

        # Use Paragraphs to allow wrapping of long fields (email, full name)
        body_style = ParagraphStyle('small', parent=styles['BodyText'], fontSize=9, leading=11)
        username_p = Paragraph(username, body_style)
        full_p = Paragraph(full or '-', body_style)
        email_p = Paragraph(email or '-', body_style)
        phone_p = Paragraph(phone or '-', body_style)
        joined_p = Paragraph(joined or '-', body_style)
        avg_p = Paragraph(str(avg_marks), body_style)
        att_p = Paragraph(f"{attendance_pct}%", body_style)

        data.append([username_p, full_p, email_p, phone_p, joined_p, avg_p, att_p])

    # Adjust column widths to give email and name more room and allow wrapping
    table = Table(data, colWidths=[0.9*inch, 1.8*inch, 2.2*inch, 1.0*inch, 0.9*inch, 0.7*inch, 0.9*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('ALIGN', (0,1), (0,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('GRID', (0,0), (-1,-1), 0.4, colors.black),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
    ]))

    story.append(table)
    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="course_{course.code}_report.pdf"'
    response.write(pdf)
    return response


@login_required
def send_message(request):
    """Send a message (Notification) to students in a course. Accepts POST: course_id, message.
    Returns JSON with sent count.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    if getattr(request.user.profile, 'role', None) != 'teacher':
        return JsonResponse({'error': 'Access denied'}, status=403)

    course_id = request.POST.get('course_id')
    message = request.POST.get('message', '').strip()
    if not message:
        return JsonResponse({'error': 'Message body is required'}, status=400)

    try:
        course = Course.objects.get(id=int(course_id), teacher=request.user.profile)
    except Exception:
        return JsonResponse({'error': 'Course not found or access denied'}, status=404)

    enrollments = Enrollment.objects.filter(course=course).select_related('student__user')
    sent = 0
    for e in enrollments:
        try:
            Notification.objects.create(user=e.student.user, title=f'Message from {request.user.get_full_name() or request.user.username}', message=message)
            sent += 1
        except Exception:
            continue

    return JsonResponse({'success': True, 'sent': sent})


@login_required
def course_detail(request, course_id):
    try:
        course = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        messages.error(request, 'Course not found')
        return redirect('teacher_dashboard')

    enrollments = Enrollment.objects.filter(course=course)
    return render(request, 'course_detail.html', {'course': course, 'enrollments': enrollments})


@login_required
def student_assignments(request):
    """List assignments visible to the logged-in student (upcoming and past)."""
    profile = getattr(request.user, 'profile', None)
    if not profile or getattr(profile, 'role', None) != 'student':
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    # Get courses the student is enrolled in
    enrollments = Enrollment.objects.filter(student=profile).select_related('course')
    courses = [e.course for e in enrollments]

    # Assignments for those courses.
    # Include assignments with no due_date (null) as 'upcoming' so teacher-created items without a date are visible.
    from django.db.models import Q
    today = date.today()
    upcoming = Assignment.objects.filter(course__in=courses).filter(Q(due_date__gte=today) | Q(due_date__isnull=True)).order_by('due_date')
    past = Assignment.objects.filter(course__in=courses, due_date__lt=today).order_by('-due_date')

    # Fetch existing submissions for quick lookup
    subs = Submission.objects.filter(student=profile, assignment__in=list(upcoming) + list(past))
    submitted_ids = set(s.assignment_id for s in subs)

    return render(request, 'assignments/student_assignments.html', {
        'upcoming': upcoming,
        'past': past,
        'submitted_ids': submitted_ids,
    })


@login_required
def student_assignment_detail(request, assignment_id):
    """Show assignment detail to a student and allow a simple submission (text-based)."""
    profile = getattr(request.user, 'profile', None)
    if not profile or getattr(profile, 'role', None) != 'student':
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    try:
        assignment = Assignment.objects.select_related('course').get(id=assignment_id)
    except Assignment.DoesNotExist:
        messages.error(request, 'Assignment not found')
        return redirect('student_assignments')

    # Permission: student must be enrolled in the course
    if not Enrollment.objects.filter(course=assignment.course, student=profile).exists():
        messages.error(request, 'Access denied')
        return redirect('role_redirect')

    # Find existing submission (if any)
    submission = Submission.objects.filter(assignment=assignment, student=profile).first()

    if request.method == 'POST':
        # Simple text submission stored in the 'feedback' field (model already present).
        submission_text = request.POST.get('submission_text', '').strip()
        submission_file = request.FILES.get('submission_file')
        if submission:
            # allow resubmission: save student's text into `text`, teacher feedback remains separate
            submission.text = submission_text
            if submission_file:
                submission.file = submission_file
            submission.submission_date = timezone.now()
            submission.save()
            messages.success(request, 'Submission updated')
        else:
            Submission.objects.create(assignment=assignment, student=profile, text=submission_text, file=submission_file)
            messages.success(request, 'Submitted successfully')
        return redirect('student_assignment_detail', assignment_id=assignment.id)

    context = {
        'assignment': assignment,
        'submission': submission,
    }
    return render(request, 'assignments/student_assignment_detail.html', context)


# Student Dashboard
@login_required
def student_dashboard(request):
    if request.user.profile.role != 'student':
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')
    
    student_profile = request.user.profile
    enrollments = Enrollment.objects.filter(student=student_profile)
    
    # attendance and performance summaries
    total_records = Attendance.objects.filter(student=student_profile).count()
    present = Attendance.objects.filter(student=student_profile, status=True).count()
    overall_attendance = (present / total_records * 100) if total_records > 0 else 0
    
    # Get live updates
    from .live_updates import get_student_live_updates
    live_updates = get_student_live_updates(student_profile)

    # Get live updates
    from .live_updates import get_student_live_updates
    live_updates = get_student_live_updates(student_profile)

    # per-course average marks for this student
    performance_labels = []
    performance_values = []
    for e in enrollments:
        performance_labels.append(e.course.name[:20])
        avg_marks = Submission.objects.filter(student=student_profile, assignment__course=e.course).aggregate(Avg('marks_obtained'))['marks_obtained__avg']
        performance_values.append(float(avg_marks) if avg_marks is not None else 0.0)

    # overall average
    avg_marks_overall = Submission.objects.filter(student=student_profile).aggregate(Avg('marks_obtained'))['marks_obtained__avg']
    avg_marks_overall = float(avg_marks_overall) if avg_marks_overall is not None else 0.0

    # Attach materials for each enrollment so student template can show them
    for e in enrollments:
        e.materials = StudyMaterial.objects.filter(course=e.course).order_by('-uploaded_at')

    # Get latest materials across all enrolled courses
    enrolled_courses = [e.course for e in enrollments]
    latest_materials = StudyMaterial.objects.filter(course__in=enrolled_courses).order_by('-uploaded_at')[:5]

    context = {
        'enrollments': enrollments,
        'total_courses': enrollments.count(),
        'overall_attendance': round(overall_attendance, 2),
        'attendance_total': total_records,
        'avg_marks_overall': round(avg_marks_overall, 2),
        'performance_labels': performance_labels,
        'performance_values': performance_values,
        'live_updates': live_updates,
        'latest_materials': latest_materials,
    }
    return render(request, 'dashboards/student_dashboard.html', context)


@login_required
def student_dashboard_adv(request):
    """Scaffolded student dashboard with quick links to attendance, fees, notifications, and reports."""
    if getattr(request.user.profile, 'role', None) != 'student':
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')
    return render(request, 'dashboards/student_dashboard_adv.html')



@login_required
def upload_material(request, course_id):
    # Only teachers for the course can upload
    if getattr(request.user.profile, 'role', None) != 'teacher':
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    try:
        course = Course.objects.get(id=course_id, teacher=request.user.profile)
    except Course.DoesNotExist:
        messages.error(request, 'Course not found or access denied')
        return redirect('teacher_dashboard')

    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description', '')
        file = request.FILES.get('file')
        if not title or not file:
            messages.error(request, 'Title and file are required')
            return redirect('upload_material', course_id=course_id)

        # Validate file type and size
        # Allowed: PDF, Word (.doc/.docx), images (PNG/JPG) and common video types
        allowed = (
            'application/pdf',
            'application/msword',
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            'image/png',
            'image/jpeg',
            'image/jpg',
            # include some common video content-types for completeness
            'video/mp4',
            'video/quicktime',
            'video/webm'
        )
        ctype = getattr(file, 'content_type', '').lower()
        # Videos get a larger size cap
        if ctype.startswith('video/'):
            max_bytes = 500 * 1024 * 1024
        else:
            max_bytes = 50 * 1024 * 1024

        if not (ctype in allowed or ctype.startswith('video/')):
            messages.error(request, 'Only PDF, Word, image, and video files are allowed (PDF, DOC/DOCX, PNG, JPG, MP4, etc.)')
            return redirect('upload_material', course_id=course_id)

        if getattr(file, 'size', 0) > max_bytes:
            messages.error(request, 'File too large (max 50MB for documents/images, 500MB for videos)')
            return redirect('upload_material', course_id=course_id)

        StudyMaterial.objects.create(course=course, uploaded_by=request.user.profile, title=title, description=description, file=file)
        messages.success(request, 'Material uploaded')
        return redirect('course_detail', course_id=course_id)

    return render(request, 'materials/upload_material.html', {'course': course})


@login_required
def teacher_manage_materials(request, course_id):
    """One-page teacher materials management: upload, list, and delete for their course."""
    # Only teachers can manage materials and must own the course
    if getattr(request.user.profile, 'role', None) != 'teacher':
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    try:
        course = Course.objects.get(id=course_id, teacher=request.user.profile)
    except Course.DoesNotExist:
        messages.error(request, 'Course not found or access denied')
        return redirect('teacher_dashboard')

    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description', '')
        file = request.FILES.get('file')
        if not title or not file:
            messages.error(request, 'Title and file are required')
            return redirect('manage_materials', course_id=course_id)

        # Validate file type and size
        allowed = (
            'application/pdf',
            'application/msword',
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            'image/png',
            'image/jpeg',
            'image/jpg',
            'video/mp4',
            'video/quicktime',
            'video/webm'
        )
        ctype = getattr(file, 'content_type', '').lower()
        if ctype.startswith('video/'):
            max_bytes = 500 * 1024 * 1024
        else:
            max_bytes = 50 * 1024 * 1024

        if not (ctype in allowed or ctype.startswith('video/')):
            messages.error(request, 'Only PDF, Word, image, and video files are allowed (PDF, DOC/DOCX, PNG, JPG, MP4, etc.)')
            return redirect('manage_materials', course_id=course_id)

        if getattr(file, 'size', 0) > max_bytes:
            messages.error(request, 'File too large (max 50MB for documents/images, 500MB for videos)')
            return redirect('manage_materials', course_id=course_id)

        StudyMaterial.objects.create(course=course, uploaded_by=request.user.profile, title=title, description=description, file=file)
        messages.success(request, 'Material uploaded')
        return redirect('manage_materials', course_id=course_id)

    materials = StudyMaterial.objects.filter(course=course).order_by('-uploaded_at')
    return render(request, 'materials/teacher_manage_materials.html', {'course': course, 'materials': materials})


@login_required
def delete_material(request, material_id):
    """Allow uploader (teacher) or superadmin to delete a material."""
    try:
        material = StudyMaterial.objects.get(id=material_id)
    except StudyMaterial.DoesNotExist:
        messages.error(request, 'Material not found')
        return redirect('role_redirect')

    user_profile = getattr(request.user, 'profile', None)
    allowed = False
    if is_admin_role(user_profile):
        allowed = True
    elif user_profile and user_profile.role == 'teacher':
        # teacher can delete if they uploaded or they teach the course
        if material.uploaded_by == user_profile or material.course.teacher == user_profile:
            allowed = True

    if not allowed:
        messages.error(request, 'Access denied')
        return redirect('role_redirect')

    course_id = material.course.id
    material.file.delete(save=False)
    material.delete()
    messages.success(request, 'Material deleted')
    # Redirect back to teacher manage page
    return redirect('manage_materials', course_id=course_id)


@login_required
def admin_upload_material(request, course_id):
    """Allow superadmins to upload study materials for any course."""
    if not is_admin_role(getattr(request.user, 'profile', None)):
        messages.error(request, 'Access denied!')
        return redirect('role_redirect')

    try:
        course = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        messages.error(request, 'Course not found')
        return redirect('admin_dashboard')

    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description', '')
        file = request.FILES.get('file')
        if not title or not file:
            messages.error(request, 'Title and file are required')
            return redirect('admin_upload_material', course_id=course_id)

        # Validate file type and size
        allowed = (
            'application/pdf',
            'application/msword',
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            'image/png',
            'image/jpeg',
            'image/jpg',
            'video/mp4',
            'video/quicktime',
            'video/webm'
        )
        ctype = getattr(file, 'content_type', '').lower()
        if ctype.startswith('video/'):
            max_bytes = 500 * 1024 * 1024
        else:
            max_bytes = 50 * 1024 * 1024

        if not (ctype in allowed or ctype.startswith('video/')):
            messages.error(request, 'Only PDF, Word, image, and video files are allowed (PDF, DOC/DOCX, PNG, JPG, MP4, etc.)')
            return redirect('admin_upload_material', course_id=course_id)

        if getattr(file, 'size', 0) > max_bytes:
            messages.error(request, 'File too large (max 50MB for documents/images, 500MB for videos)')
            return redirect('admin_upload_material', course_id=course_id)

        StudyMaterial.objects.create(course=course, uploaded_by=request.user.profile, title=title, description=description, file=file)
        messages.success(request, 'Material uploaded')
        return redirect('admin_dashboard')

    return render(request, 'materials/upload_material.html', {'course': course, 'admin_upload': True})


@login_required
def materials_list(request, course_id):
    # Students and teachers can view materials for a course they are enrolled/teaching
    try:
        course = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        messages.error(request, 'Course not found')
        return redirect('role_redirect')

    # basic permission: student must be enrolled or teacher must be owner
    if request.user.profile.role == 'student':
        if not Enrollment.objects.filter(course=course, student=request.user.profile).exists():
            messages.error(request, 'Access denied')
            return redirect('role_redirect')

    materials = StudyMaterial.objects.filter(course=course).order_by('-uploaded_at')
    return render(request, 'materials/materials_list.html', {'course': course, 'materials': materials})


@login_required
def download_material(request, material_id):
    try:
        material = StudyMaterial.objects.get(id=material_id)
    except StudyMaterial.DoesNotExist:
        messages.error(request, 'Material not found')
        return redirect('role_redirect')

    # permission checks
    course = material.course
    if request.user.profile.role == 'student':
        if not Enrollment.objects.filter(course=course, student=request.user.profile).exists():
            messages.error(request, 'Access denied')
            return redirect('role_redirect')

    response = HttpResponse(material.file, content_type='application/octet-stream')
    response['Content-Disposition'] = f'attachment; filename="{material.file.name.split("/")[-1]}"'
    return response


@login_required
def view_material(request, material_id):
    """Serve material inline for PDFs (so the browser will render it) or fall back to attachment.

    Students must be enrolled to view materials for their course.
    """
    try:
        material = StudyMaterial.objects.get(id=material_id)
    except StudyMaterial.DoesNotExist:
        messages.error(request, 'Material not found')
        return redirect('role_redirect')

    course = material.course
    if request.user.profile.role == 'student':
        if not Enrollment.objects.filter(course=course, student=request.user.profile).exists():
            messages.error(request, 'Access denied')
            return redirect('role_redirect')

    # Determine if file is a PDF by extension
    name = material.file.name.lower()
    if name.endswith('.pdf'):
        content_type = 'application/pdf'
        disposition = 'inline'
    else:
        content_type = 'application/octet-stream'
        disposition = 'attachment'

    response = HttpResponse(material.file, content_type=content_type)
    response['Content-Disposition'] = f'{disposition}; filename="{material.file.name.split("/")[-1]}"'
    return response


@login_required
def submit_feedback(request, course_id=None):
    if request.method == 'POST':
        message = request.POST.get('message')
        course = None
        if course_id:
            try:
                course = Course.objects.get(id=course_id)
            except Course.DoesNotExist:
                course = None

        Feedback.objects.create(student=request.user.profile, course=course, message=message)
        messages.success(request, 'Thanks for your feedback')
        return redirect('student_dashboard')

    return render(request, 'feedback/submit_feedback.html', {'course_id': course_id})

